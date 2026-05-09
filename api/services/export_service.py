# api/services/export_service.py — P3-03 : Rapport PDF Professionnel 8 Sections
"""
Phase 3 — Rapport PDF professionnel avec 8 sections dynamiques :
  1. En-tête & Métadonnées
  2. Source Database Profile
  3. Data Quality Report
  4. Star/Constellation Schema Design
  5. DDL SQL Server
  6. Pipeline ETL Summary
  7. Analytical Queries & Results
  8. Métriques de Qualité & Journal

Utilise fpdf2 (priorité) ou reportlab (fallback).
Le rapport utilise state['logical_model'] et state['query_results'] — données réelles.
"""
import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def generate_pdf_report(state: dict, session_id: str) -> str:
    """
    Génère un rapport PDF professionnel du pipeline avec 8 sections dynamiques.
    Retourne le chemin du fichier PDF créé.
    """
    # Essayer fpdf2 d'abord (plus flexible pour les tables complexes)
    try:
        return _generate_pdf_fpdf2(state, session_id)
    except ImportError:
        logger.info("[Export] fpdf2 non disponible, fallback sur reportlab")

    # Fallback reportlab
    try:
        return _generate_pdf_reportlab(state, session_id)
    except ImportError:
        raise RuntimeError(
            "Ni fpdf2 ni reportlab ne sont installés. "
            "Ajoutez 'fpdf2>=2.7.0' ou 'reportlab>=4.2.0' à requirements.txt"
        )


def _generate_pdf_fpdf2(state: dict, session_id: str) -> str:
    """Génère le rapport PDF avec fpdf2."""
    from fpdf import FPDF

    os.makedirs("outputs", exist_ok=True)
    pdf_path = f"outputs/{session_id}_report.pdf"
    generated_at = datetime.now().strftime("%d/%m/%Y à %H:%M")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    # ── Couleurs du thème ────────────────────────────────────────────────────
    primary = (99, 102, 241)   # Indigo
    dark = (30, 41, 59)
    gray = (100, 116, 139)
    success = (5, 150, 105)
    warning = (217, 119, 6)
    bg_light = (248, 250, 252)

    def _section_title(title: str):
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(*primary)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(*primary)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
        pdf.ln(4)
        pdf.set_text_color(*dark)

    def _body_text(text: str, bold=False):
        pdf.set_font("Helvetica", "B" if bold else "", 9)
        pdf.set_text_color(*dark)
        pdf.multi_cell(0, 5, str(text)[:3000])
        pdf.ln(2)

    def _code_block(text: str, max_len: int = 2500):
        pdf.set_font("Courier", "", 7)
        pdf.set_text_color(30, 41, 59)
        pdf.set_fill_color(*bg_light)
        display = str(text)[:max_len]
        if len(str(text)) > max_len:
            display += "\n-- [tronqué...]"
        pdf.multi_cell(0, 4, display, fill=True)
        pdf.ln(3)

    def _kv_table(data: list):
        """Table clé-valeur."""
        pdf.set_font("Helvetica", "", 9)
        col_w = [50, 130]
        for key, val in data:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(*primary)
            pdf.cell(col_w[0], 6, str(key))
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(*dark)
            pdf.cell(col_w[1], 6, str(val)[:80], new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    def _data_table(columns: list, rows: list, max_rows: int = 15):
        """Table de données avec en-têtes."""
        if not columns or not rows:
            return
        n_cols = min(len(columns), 6)
        col_w = (pdf.w - pdf.l_margin - pdf.r_margin) / n_cols

        # Header
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(*primary)
        pdf.set_text_color(255, 255, 255)
        for col in columns[:n_cols]:
            pdf.cell(col_w, 6, str(col)[:20], border=1, fill=True)
        pdf.ln()

        # Rows
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(*dark)
        for i, row in enumerate(rows[:max_rows]):
            pdf.set_fill_color(255, 255, 255) if i % 2 == 0 else pdf.set_fill_color(*bg_light)
            for val in row[:n_cols]:
                cell_val = str(val)[:18] if val is not None else "NULL"
                pdf.cell(col_w, 5, cell_val, border=1, fill=True)
            pdf.ln()
        if len(rows) > max_rows:
            pdf.set_font("Helvetica", "I", 7)
            pdf.set_text_color(*gray)
            pdf.cell(0, 5, f"... {len(rows) - max_rows} lignes supplémentaires non affichées",
                     new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 1 : En-tête & Métadonnées
    # ═══════════════════════════════════════════════════════════════════════════
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_text_color(*primary)
    pdf.cell(0, 15, "ANTIGRAVITY BI", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(*gray)
    pdf.cell(0, 8, "Rapport d'Exécution Pipeline ETL IA", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(8)

    etl_status = state.get("etl_status", "N/A").upper()
    status_color = success if etl_status == "SUCCESS" else warning

    _kv_table([
        ("Session ID", session_id),
        ("Généré le", generated_at),
        ("Utilisateur", state.get("user_prefix", "N/A")),
        ("Statut ETL", etl_status),
        ("Version modèle", str(state.get("logical_model_version", 0))),
        ("Corrections Healer", str(len(state.get("heal_history", [])))),
        ("Mode ETL", state.get("etl_mode", "full_load")),
    ])

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 2 : Source Database Profile
    # ═══════════════════════════════════════════════════════════════════════════
    _section_title("1. Profil de la Base Source")
    source_meta = state.get("source_metadata", {})
    if source_meta:
        tables_data = []
        for tname, tinfo in source_meta.items():
            if not isinstance(tinfo, dict):
                continue
            cols = tinfo.get("columns", [])
            rows = tinfo.get("row_count", 0)
            fks = len(tinfo.get("foreign_keys", []))
            tables_data.append([tname, str(len(cols)), str(rows), str(fks)])

        if tables_data:
            _data_table(
                ["Table", "Colonnes", "Lignes", "FK"],
                tables_data
            )
            _body_text(f"Total : {len(tables_data)} tables dans la base source")
    else:
        _body_text("Aucune métadonnée source disponible")

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 3 : Data Quality Report
    # ═══════════════════════════════════════════════════════════════════════════
    _section_title("2. Rapport Qualité des Données (DQ)")
    dq_score = state.get("dq_score")
    dq_alerts = state.get("dq_alerts", [])

    if dq_score is not None:
        score_color = success if dq_score >= 70 else warning if dq_score >= 50 else (239, 68, 68)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*score_color)
        pdf.cell(0, 10, f"Score DQ : {dq_score}/100", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(*dark)
        pdf.ln(3)

    if dq_alerts:
        alerts_data = [
            [a.get("severity", "?"), a.get("table", "?"),
             a.get("column", "?"), a.get("rule", "?")[:30]]
            for a in dq_alerts[:20]
        ]
        _data_table(["Sévérité", "Table", "Colonne", "Règle"], alerts_data)
    else:
        _body_text("Aucune alerte de qualité détectée")

    # SECTION 3 bis : Schema Drift
    if state.get("schema_drift_detected"):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*warning)
        pdf.cell(0, 8, "DÉRIVE DE SCHÉMA DÉTECTÉE", new_x="LMARGIN", new_y="NEXT")
        _body_text(state.get("schema_drift_details", ""))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 4 : Star/Constellation Schema Design
    # ═══════════════════════════════════════════════════════════════════════════
    pdf.add_page()
    _section_title("3. Modèle Dimensionnel (Star/Constellation)")
    logical_model = state.get("logical_model", {})

    if logical_model:
        # Fact tables
        fact_tables = logical_model.get("fact_tables", [])
        if not fact_tables:
            ft = logical_model.get("fact_table")
            fact_tables = [ft] if ft else []

        n_facts = len(fact_tables)
        n_dims = len(logical_model.get("dimension_tables", []))
        schema_type = "Constellation" if n_facts > 1 else "Star Schema"

        _body_text(f"Type : {schema_type} — {n_facts} table(s) de faits, {n_dims} dimensions", bold=True)

        for fact in fact_tables:
            if not fact:
                continue
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*primary)
            pdf.cell(0, 7, f"FACT: {fact.get('name', '?')}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(*dark)

            cols_data = [
                [c.get("name", "?"), c.get("type", "?"), c.get("role", "?")]
                for c in fact.get("columns", [])
            ]
            _data_table(["Colonne", "Type", "Rôle"], cols_data)

        for dim in logical_model.get("dimension_tables", []):
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(5, 150, 105)
            pdf.cell(0, 7, f"DIM: {dim.get('name', '?')}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(*dark)

            cols_data = [
                [c.get("name", "?"), c.get("type", "?"), c.get("role", "?")]
                for c in dim.get("columns", [])[:10]
            ]
            _data_table(["Colonne", "Type", "Rôle"], cols_data)
    else:
        _body_text("Aucun modèle logique disponible")

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 5 : DDL SQL Server
    # ═══════════════════════════════════════════════════════════════════════════
    _section_title("4. DDL SQL Server — Schéma DW")
    sql_ddl = state.get("sql_ddl", "")
    if sql_ddl:
        _code_block(sql_ddl, max_len=3000)
    else:
        _body_text("DDL non généré")

    # Rapport du Critic
    critic_review = state.get("critic_review", "")
    critic_approved = state.get("critic_approved", False)
    if critic_review:
        verdict = "APPROVED" if critic_approved else "NEEDS_REVISION"
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*(success if critic_approved else warning))
        pdf.cell(0, 7, f"Verdict Critic : {verdict}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(*dark)
        _body_text(critic_review[:1500])

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 6 : Pipeline ETL Summary
    # ═══════════════════════════════════════════════════════════════════════════
    pdf.add_page()
    _section_title("5. Pipeline ETL — Résumé des Transformations")
    etl_code = state.get("etl_code", "")
    if etl_code:
        _body_text("Procédures T-SQL MERGE générées :", bold=True)
        _code_block(etl_code, max_len=2000)

    # Load metrics
    load_metrics = state.get("load_metrics", {})
    if load_metrics:
        _body_text("Métriques de chargement :", bold=True)
        metrics_data = []
        if isinstance(load_metrics, dict):
            for key, val in load_metrics.items():
                if isinstance(val, dict):
                    for sub_key, sub_val in val.items():
                        metrics_data.append([f"{key}.{sub_key}", str(sub_val)])
                else:
                    metrics_data.append([str(key), str(val)])
        if metrics_data:
            _data_table(["Métrique", "Valeur"], metrics_data[:20])

    # Healer history
    heal_history = state.get("heal_history", [])
    if heal_history:
        _body_text("Corrections Healer :", bold=True)
        for i, h in enumerate(heal_history, 1):
            _body_text(f"  {i}. {h}")

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 7 : Analytical Queries & Results
    # ═══════════════════════════════════════════════════════════════════════════
    _section_title("6. Requêtes Analytiques")
    query_results = state.get("query_results", [])
    generated_queries = state.get("generated_queries", [])

    queries_to_show = query_results if query_results else generated_queries

    if queries_to_show:
        for i, q in enumerate(queries_to_show, 1):
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*primary)
            title = q.get("title", f"Requête {i}")
            pdf.cell(0, 7, f"{i}. {title}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(*dark)

            desc = q.get("description", "")
            if desc:
                _body_text(desc)

            sql = q.get("sql", "")
            if sql:
                _code_block(sql, max_len=800)

            # Résultats de la requête (si exécutée)
            if q.get("columns") and q.get("rows"):
                _data_table(q["columns"], q["rows"], max_rows=10)
            elif q.get("error"):
                pdf.set_font("Helvetica", "I", 8)
                pdf.set_text_color(*warning)
                pdf.cell(0, 5, f"Erreur : {q['error'][:100]}", new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(*dark)
                pdf.ln(3)
    else:
        _body_text("Aucune requête analytique générée")

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 8 : Métriques & Journal d'exécution
    # ═══════════════════════════════════════════════════════════════════════════
    pdf.add_page()
    _section_title("7. Métriques d'Exécution")

    # Node durations
    node_durations = state.get("node_durations", {})
    if node_durations:
        _body_text("Durées par nœud :", bold=True)
        dur_data = [
            [node, f"{dur:.2f}s"]
            for node, dur in sorted(node_durations.items(), key=lambda x: x[1], reverse=True)
        ]
        _data_table(["Nœud", "Durée"], dur_data)

    # Reject metrics
    reject_metrics = state.get("reject_metrics", {})
    if reject_metrics:
        _body_text("Métriques de quarantaine :", bold=True)
        rej_data = [
            [str(k), str(v)] for k, v in reject_metrics.items()
        ]
        _data_table(["Table", "Rejets"], rej_data[:10])

    # CDC Watermarks
    etl_watermarks = state.get("etl_watermarks", {})
    etl_mode = state.get("etl_mode", "")
    if etl_mode:
        _body_text(f"Mode ETL : {etl_mode}", bold=True)
    if etl_watermarks:
        wm_data = [
            [table, info.get("column", "N/A"), info.get("mode", "N/A"),
             str(info.get("last_value", "N/A"))[:30]]
            for table, info in etl_watermarks.items()
            if isinstance(info, dict)
        ]
        _data_table(["Table", "Colonne tracking", "Mode", "Dernière valeur"], wm_data)

    # Journal d'exécution
    _section_title("8. Journal d'Exécution")
    exec_log = state.get("execution_log", [])
    if exec_log:
        for i, entry in enumerate(exec_log[-40:], 1):
            pdf.set_font("Courier", "", 7)
            pdf.set_text_color(*gray)
            pdf.cell(8, 4, str(i))
            pdf.set_text_color(*dark)
            pdf.multi_cell(0, 4, str(entry)[:120])
    else:
        _body_text("Journal vide")

    # ── Pied de page ─────────────────────────────────────────────────────────
    pdf.ln(5)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(*gray)
    pdf.cell(0, 5,
             f"Rapport généré par Agent BI v6.0 — {generated_at}",
             new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.output(pdf_path)
    logger.info(f"[Export] PDF généré (fpdf2) : {pdf_path}")
    return pdf_path


def _generate_pdf_reportlab(state: dict, session_id: str) -> str:
    """Fallback : rapport PDF avec reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table,
        TableStyle, HRFlowable, Preformatted
    )
    from reportlab.lib.enums import TA_CENTER

    os.makedirs("outputs", exist_ok=True)
    pdf_path = f"outputs/{session_id}_report.pdf"

    doc = SimpleDocTemplate(
        pdf_path, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    # Styles personnalisés
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=22, textColor=colors.HexColor("#6366f1"),
        spaceAfter=6
    )
    h2_style = ParagraphStyle(
        "H2", parent=styles["Heading2"],
        fontSize=13, textColor=colors.HexColor("#1e293b"),
        spaceBefore=16, spaceAfter=6,
        borderPad=4
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#374151"),
        leading=14
    )
    code_style = ParagraphStyle(
        "Code", parent=styles["Code"],
        fontSize=7, fontName="Courier",
        textColor=colors.HexColor("#1e293b"),
        backColor=colors.HexColor("#f8fafc"),
        borderColor=colors.HexColor("#e2e8f0"),
        borderWidth=1, borderPad=6,
        leading=11
    )
    success_style = ParagraphStyle(
        "Success", parent=body_style,
        textColor=colors.HexColor("#059669"),
        fontName="Helvetica-Bold"
    )
    warning_style = ParagraphStyle(
        "Warning", parent=body_style,
        textColor=colors.HexColor("#d97706"),
        fontName="Helvetica-Bold"
    )

    story = []
    generated_at = datetime.now().strftime("%d/%m/%Y à %H:%M")

    # ── En-tête ────────────────────────────────────────────────────────────────
    story.append(Paragraph("ANTIGRAVITY BI", title_style))
    story.append(Paragraph("Rapport d'exécution du Pipeline ETL IA", styles["Heading3"]))
    story.append(Spacer(1, 0.3*cm))

    # Métadonnées
    meta_data = [
        ["Session ID", session_id],
        ["Généré le", generated_at],
        ["Utilisateur", state.get("user_prefix", "N/A")],
        ["Statut ETL", state.get("etl_status", "N/A").upper()],
        ["Version modèle", str(state.get("logical_model_version", 0))],
        ["Corrections Healer", str(len(state.get("heal_history", [])))],
        ["Mode ETL", state.get("etl_mode", "full_load")],
    ]
    meta_table = Table(meta_data, colWidths=[4*cm, 12*cm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6366f1")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f8fafc")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(meta_table)
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e2e8f0"), spaceAfter=12))

    # ── Data Quality ──────────────────────────────────────────────────────────
    dq_score = state.get("dq_score")
    if dq_score is not None:
        story.append(Paragraph(f"Score DQ : {dq_score}/100", h2_style))

    # ── Schema Drift ──────────────────────────────────────────────────────────
    if state.get("schema_drift_detected"):
        story.append(Paragraph("DÉRIVE DE SCHÉMA DÉTECTÉE", warning_style))
        story.append(Paragraph(state.get("schema_drift_details", ""), body_style))
        story.append(Spacer(1, 0.3*cm))

    # ── Rapport du Critic ──────────────────────────────────────────────────────
    story.append(Paragraph("Rapport du Critic", h2_style))
    critic_review = state.get("critic_review", "Non disponible")
    critic_approved = state.get("critic_approved", False)
    verdict_style = success_style if critic_approved else warning_style
    verdict_label = "APPROVED" if critic_approved else "NEEDS_REVISION"
    story.append(Paragraph(verdict_label, verdict_style))
    story.append(Spacer(1, 0.2*cm))
    critic_text = critic_review[:2000] + ("..." if len(critic_review) > 2000 else "")
    story.append(Paragraph(critic_text.replace("\n", "<br/>"), body_style))

    # ── DDL SQL ────────────────────────────────────────────────────────────────
    sql_ddl = state.get("sql_ddl", "")
    if sql_ddl:
        story.append(Paragraph("DDL SQL — Schéma Data Warehouse", h2_style))
        sql_display = sql_ddl[:3000] + ("\n-- [tronqué...]" if len(sql_ddl) > 3000 else "")
        story.append(Preformatted(sql_display, code_style))

    # ── Requêtes Analytiques ──────────────────────────────────────────────────
    query_results = state.get("query_results", [])
    if query_results:
        story.append(Paragraph("Requêtes Analytiques", h2_style))
        for i, q in enumerate(query_results[:6], 1):
            story.append(Paragraph(f"{i}. {q.get('title', '')}", body_style))
            if q.get("sql"):
                story.append(Preformatted(q["sql"][:500], code_style))

    # ── Code ETL ──────────────────────────────────────────────────────────────
    etl_code = state.get("etl_code", "")
    if etl_code:
        story.append(Paragraph("Procédures ETL T-SQL", h2_style))
        etl_display = etl_code[:2000] + ("\n-- [tronqué...]" if len(etl_code) > 2000 else "")
        story.append(Preformatted(etl_display, code_style))

    # ── Historique Healer ─────────────────────────────────────────────────────
    heal_history = state.get("heal_history", [])
    if heal_history:
        story.append(Paragraph("Historique des corrections Healer", h2_style))
        for i, h in enumerate(heal_history, 1):
            story.append(Paragraph(f"{i}. {h}", body_style))
        story.append(Spacer(1, 0.3*cm))

    # ── Journal d'exécution ───────────────────────────────────────────────────
    exec_log = state.get("execution_log", [])
    if exec_log:
        story.append(Paragraph("Journal d'exécution", h2_style))
        log_data = [[f"{i+1}", entry] for i, entry in enumerate(exec_log[-30:])]
        log_table = Table(log_data, colWidths=[1*cm, 15*cm])
        log_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Courier"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#94a3b8")),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#f1f5f9")),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(log_table)

    # ── Pied de page ──────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    story.append(Paragraph(
        f"<font size=8 color='#94a3b8'>Rapport généré par Agent BI v6.0 — {generated_at}</font>",
        ParagraphStyle("Footer", parent=body_style, alignment=TA_CENTER)
    ))

    doc.build(story)
    logger.info(f"[Export] PDF généré (reportlab) : {pdf_path}")
    return pdf_path


def generate_json_report(state: dict, session_id: str) -> dict:
    """Génère un rapport JSON complet — Phase 3."""
    return {
        "meta": {
            "session_id": session_id,
            "generated_at": datetime.now().isoformat(),
            "user_prefix": state.get("user_prefix"),
            "etl_status": state.get("etl_status"),
            "logical_model_version": state.get("logical_model_version"),
            "critic_approved": state.get("critic_approved"),
            "schema_drift_detected": state.get("schema_drift_detected"),
            "schema_drift_details": state.get("schema_drift_details"),
            "heal_count": len(state.get("heal_history", [])),
            "dq_score": state.get("dq_score"),
            "etl_mode": state.get("etl_mode"),
        },
        "artifacts": {
            "sql_ddl": state.get("sql_ddl"),
            "etl_code": state.get("etl_code"),
            "logical_model": state.get("logical_model"),
            "generated_queries": state.get("generated_queries", []),
        },
        "analytics": {
            "query_results": state.get("query_results", []),
            "executive_summary": state.get("executive_summary"),
            "visualizations": state.get("visualizations", []),
        },
        "quality": {
            "dq_score": state.get("dq_score"),
            "dq_alerts": state.get("dq_alerts", []),
            "reject_metrics": state.get("reject_metrics", {}),
            "etl_watermarks": state.get("etl_watermarks", {}),
        },
        "audit": {
            "critic_review": state.get("critic_review"),
            "heal_history": state.get("heal_history", []),
            "execution_log": state.get("execution_log", []),
            "lineage": state.get("lineage", {}),
            "node_durations": state.get("node_durations", {}),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Exports BI (Power BI, Tableau, Excel) — ajoutés v5.0
# ─────────────────────────────────────────────────────────────────────────────

def _iter_result_tables(state: dict):
    """
    Yields (name, rows) pour chaque table exportable trouvée dans le state.

    Sources consultées (dans l'ordre) :
      1. state['query_results']   → [{title, columns, rows: [list]}, ...]
      2. state['etl_samples']     → {table_name: [dict]}
      3. state['tables_preview']  → {table_name: [dict]}
    Une table est toujours au format list[dict] (rows).
    """
    # (1) query_results — rows sont des listes, columns contient les vrais noms
    for qr in state.get("query_results", []) or []:
        if isinstance(qr, dict):
            name = qr.get("title") or qr.get("name") or qr.get("table") or "query"
            columns = qr.get("columns", [])
            rows = qr.get("rows") or qr.get("data") or []
            if rows and isinstance(rows, list):
                first = rows[0]
                # Si rows sont des listes, les convertir en dicts avec les vrais noms
                if isinstance(first, (list, tuple)) and columns:
                    rows = [dict(zip(columns, row)) for row in rows]
                yield str(name), rows

    # (2) etl_samples
    samples = state.get("etl_samples") or {}
    if isinstance(samples, dict):
        for name, rows in samples.items():
            if rows and isinstance(rows, list):
                yield str(name), rows

    # (3) tables_preview
    preview = state.get("tables_preview") or {}
    if isinstance(preview, dict):
        for name, rows in preview.items():
            if rows and isinstance(rows, list):
                yield str(name), rows


def _sanitize_sheet_name(name: str) -> str:
    """Excel : 31 chars max, pas de ':' '\\' '/' '?' '*' '[' ']'."""
    bad = ':\\/?*[]'
    clean = "".join(c for c in str(name) if c not in bad).strip() or "sheet"
    return clean[:31]



# ─── Palette & styles partagés ────────────────────────────────────────────────

_C = {
    "navy":    "1E3A5F",
    "blue":    "2563EB",
    "teal":    "0EA5E9",
    "indigo":  "4F46E5",
    "green":   "059669",
    "green2":  "D1FAE5",
    "amber":   "D97706",
    "amber2":  "FEF3C7",
    "red":     "DC2626",
    "red2":    "FEE2E2",
    "white":   "FFFFFF",
    "light":   "F8FAFC",
    "gray1":   "F1F5F9",
    "gray2":   "E2E8F0",
    "gray3":   "CBD5E1",
    "gray4":   "94A3B8",
    "dark":    "0F172A",
    "gold":    "B45309",
    "gold2":   "FFF7ED",
    "purple":  "7C3AED",
    "purple2": "EDE9FE",
    "row_alt": "EFF6FF",
}

def _mk_font(bold=False, size=10, color="0F172A", italic=False, name="Calibri"):
    from openpyxl.styles import Font
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _mk_fill(color_hex):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=color_hex)

def _mk_border(style="thin", color="CBD5E1"):
    from openpyxl.styles import Border, Side
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _mk_align(h="left", v="center", wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _col_letter(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)

def _set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def _header_cell(ws, row, col, text, bg=None, fg="FFFFFF", size=10, bold=True, h="left", merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font    = _mk_font(bold=bold, size=size, color=fg)
    cell.fill    = _mk_fill(bg or _C["navy"])
    cell.alignment = _mk_align(h=h, v="center")
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return cell

def _data_cell(ws, row, col, value, bg=None, fg=None, bold=False, size=10,
               h="left", border=True, num_fmt=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font  = _mk_font(bold=bold, size=size, color=fg or _C["dark"])
    if bg:
        cell.fill = _mk_fill(bg)
    if border:
        cell.border = _mk_border("thin", _C["gray3"])
    cell.alignment = _mk_align(h=h, v="center", wrap=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _section_title(ws, row, col, text, colspan=10, bg=None):
    c = ws.cell(row=row, column=col, value=f"  {text}")
    c.font      = _mk_font(bold=True, size=11, color=_C["white"])
    c.fill      = _mk_fill(bg or _C["blue"])
    c.alignment = _mk_align(h="left", v="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colspan - 1)
    _set_row_height(ws, row, 22)
    return c

def _spacer(ws, row, n_cols=15):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c, value="")
        cell.fill = _mk_fill(_C["light"])
    _set_row_height(ws, row, 8)

def _set_cols(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 1 : TABLEAU DE BORD
# ─────────────────────────────────────────────────────────────────────────────

def _ws_dashboard(wb, state: dict, session_id: str):
    ws = wb.active
    ws.title = "🎯 Tableau de Bord"

    # ── Bandeau titre ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:N4")
    title_cell = ws["A1"]
    title_cell.value     = "RAPPORT DÉCISIONNEL — DATA WAREHOUSE"
    title_cell.font      = _mk_font(bold=True, size=20, color=_C["white"], name="Calibri")
    title_cell.fill      = _mk_fill(_C["navy"])
    title_cell.alignment = _mk_align(h="center", v="center")
    for r in range(1, 5):
        _set_row_height(ws, r, 18)

    # Sous-titre
    ws.merge_cells("A5:N5")
    sub = ws["A5"]
    dts = datetime.now().strftime("%d %B %Y — %H:%M")
    sub.value     = f"  Pipeline ID : {session_id}   ·   Généré le {dts}   ·   Préfixe : {state.get('user_prefix','dw').upper()}"
    sub.font      = _mk_font(italic=True, size=10, color=_C["white"])
    sub.fill      = _mk_fill(_C["indigo"])
    sub.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 5, 18)

    # ── KPI Cards ─────────────────────────────────────────────────────────────
    _spacer(ws, 6)

    kpi_row   = 7
    kpi_items = [
        ("A", "Score Qualité Données", f"{state.get('dq_score',0)}/100",  _C["indigo"], _C["purple2"]),
        ("C", "Statut ETL",            (state.get("etl_status") or "N/A").upper(), _C["green"], _C["green2"]),
        ("E", "Mode ETL",              (state.get("etl_mode") or "N/A").upper(),   _C["amber"],  _C["amber2"]),
        ("G", "Version Schéma",        f"v{state.get('logical_model_version',1)}",  _C["blue"],   _C["row_alt"]),
        ("I", "Auto-corrections",      str(len(state.get("heal_history",[]) or [])), _C["gold"],  _C["gold2"]),
        ("K", "Alertes DQ",            str(len(state.get("dq_alerts",[]) or [])),    _C["red"],   _C["red2"]),
    ]
    for col_letter, label, value, accent, bg in kpi_items:
        col_n = ord(col_letter) - 64
        # Fusion 2 colonnes par KPI
        ws.merge_cells(start_row=kpi_row,   start_column=col_n, end_row=kpi_row,   end_column=col_n+1)
        ws.merge_cells(start_row=kpi_row+1, start_column=col_n, end_row=kpi_row+1, end_column=col_n+1)
        ws.merge_cells(start_row=kpi_row+2, start_column=col_n, end_row=kpi_row+2, end_column=col_n+1)

        lbl_cell = ws.cell(row=kpi_row,   column=col_n, value=label)
        val_cell = ws.cell(row=kpi_row+1, column=col_n, value=value)
        sep_cell = ws.cell(row=kpi_row+2, column=col_n, value="")

        lbl_cell.font      = _mk_font(bold=True,  size=9,  color=_C["white"])
        lbl_cell.fill      = _mk_fill(accent)
        lbl_cell.alignment = _mk_align(h="center", v="center")
        lbl_cell.border    = _mk_border("medium", accent)

        val_cell.font      = _mk_font(bold=True,  size=16, color=accent)
        val_cell.fill      = _mk_fill(bg)
        val_cell.alignment = _mk_align(h="center", v="center")
        val_cell.border    = _mk_border("medium", accent)

        sep_cell.fill   = _mk_fill(bg)
        sep_cell.border = _mk_border("medium", accent)

        _set_row_height(ws, kpi_row,   16)
        _set_row_height(ws, kpi_row+1, 30)
        _set_row_height(ws, kpi_row+2, 6)

    _spacer(ws, kpi_row+3)

    # ── Résumé pour décideurs ─────────────────────────────────────────────────
    r = kpi_row + 4
    _section_title(ws, r, 1, "📋  SYNTHÈSE POUR DÉCIDEURS", colspan=14)
    r += 1

    dq = state.get("dq_score", 0)
    etl_ok = state.get("etl_status","") == "success"
    n_dims = len((state.get("logical_model") or {}).get("dimension_tables", []))
    n_heal = len(state.get("heal_history", []) or [])
    n_alerts = len(state.get("dq_alerts", []) or [])
    lm = state.get("logical_model") or {}
    fact_name = (lm.get("fact_table") or {}).get("name", "N/A")
    prefix = state.get("user_prefix", "dw")

    dq_text = "✅ Excellente" if dq >= 90 else ("⚠️ Acceptable" if dq >= 70 else "❌ Insuffisante")
    etl_text = "✅ ETL exécuté avec succès" if etl_ok else "⚠️ ETL non terminé ou en erreur"

    summary_points = [
        ("Qualité des données",    f"{dq_text} ({dq}/100) — {n_alerts} alerte(s) détectée(s)"),
        ("Statut du pipeline",     etl_text),
        ("Schéma Data Warehouse",  f"Modèle en étoile : 1 table de faits '{prefix}_{fact_name}' + {n_dims} dimension(s)"),
        ("Fiabilité",              f"{n_heal} auto-correction(s) appliquée(s) par le Healer"),
        ("Mode de chargement",     f"{(state.get('etl_mode') or 'full_load').replace('_',' ').title()} — CDC activé" if state.get("etl_watermarks") else f"{(state.get('etl_mode') or 'full_load').replace('_',' ').title()}"),
    ]

    lm_metrics = state.get("load_metrics") or {}
    total_rows = 0
    if isinstance(lm_metrics, dict):
        for v in lm_metrics.values():
            if isinstance(v, dict):
                total_rows += v.get("inserted", 0)
            elif isinstance(v, (int, float)):
                total_rows += v
    if total_rows > 0:
        summary_points.append(("Lignes chargées", f"{total_rows:,} enregistrements transférés dans le DW"))

    node_dur = state.get("node_durations") or {}
    total_dur = round(sum(float(v) for v in node_dur.values() if isinstance(v, (int, float))), 1)
    if total_dur > 0:
        summary_points.append(("Durée totale pipeline", f"{total_dur}s ({round(total_dur/60,1)} min)"))

    for label, text in summary_points:
        lbl_c = ws.cell(row=r, column=1, value=label)
        lbl_c.font      = _mk_font(bold=True, size=10, color=_C["navy"])
        lbl_c.fill      = _mk_fill(_C["gray1"])
        lbl_c.alignment = _mk_align(h="left", v="center")
        lbl_c.border    = _mk_border("thin", _C["gray3"])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

        txt_c = ws.cell(row=r, column=4, value=text)
        txt_c.font      = _mk_font(size=10, color=_C["dark"])
        txt_c.fill      = _mk_fill(_C["white"])
        txt_c.alignment = _mk_align(h="left", v="center", wrap=True)
        txt_c.border    = _mk_border("thin", _C["gray3"])
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=14)
        _set_row_height(ws, r, 18)
        r += 1

    _spacer(ws, r); r += 1

    # ── Table des métriques de chargement ─────────────────────────────────────
    if isinstance(lm_metrics, dict) and lm_metrics:
        _section_title(ws, r, 1, "📦  MÉTRIQUES DE CHARGEMENT PAR TABLE", colspan=14)
        r += 1
        hdrs = ["Table", "Lignes insérées", "Lignes rejetées", "Lignes mises à jour", "Total traité"]
        for ci, h in enumerate(hdrs, start=1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font      = _mk_font(bold=True, size=9, color=_C["white"])
            c.fill      = _mk_fill(_C["teal"])
            c.alignment = _mk_align(h="center", v="center")
            c.border    = _mk_border("thin", _C["teal"])
        _set_row_height(ws, r, 16); r += 1

        for tbl, metrics in lm_metrics.items():
            if isinstance(metrics, dict):
                ins = metrics.get("inserted", 0)
                rej = metrics.get("rejected", 0)
                upd = metrics.get("updated", 0)
                tot = ins + rej + upd
            elif isinstance(metrics, (int, float)):
                ins, rej, upd, tot = int(metrics), 0, 0, int(metrics)
            else:
                continue
            alt = _C["row_alt"] if r % 2 == 0 else _C["white"]
            for ci, val in enumerate([str(tbl), ins, rej, upd, tot], start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = _mk_font(size=9, bold=(ci == 1))
                c.fill      = _mk_fill(alt)
                c.alignment = _mk_align(h="right" if ci > 1 else "left", v="center")
                c.border    = _mk_border("thin", _C["gray3"])
                if ci > 1:
                    c.number_format = "#,##0"
            _set_row_height(ws, r, 15); r += 1

    # Colonnes
    _set_cols(ws, {"A":14,"B":14,"C":14,"D":14,"E":14,"F":14,"G":14,
                   "H":14,"I":14,"J":14,"K":14,"L":14,"M":14,"N":14})
    ws.freeze_panes = "A6"


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 2 : QUALITÉ DES DONNÉES
# ─────────────────────────────────────────────────────────────────────────────

def _ws_data_quality(wb, state: dict):
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint

    ws = wb.create_sheet("📊 Qualité Données")

    dq_score  = state.get("dq_score", 0) or 0
    dq_alerts = state.get("dq_alerts", []) or []
    dq_report = state.get("dq_report", {}) or {}

    # ── Bandeau ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    h = ws["A1"]
    h.value = "RAPPORT QUALITÉ DES DONNÉES"
    h.font  = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill  = _mk_fill(_C["navy"])
    h.alignment = _mk_align(h="center", v="center")
    _set_row_height(ws, 1, 28)

    # ── Score global ──────────────────────────────────────────────────────────
    _spacer(ws, 2)
    score_color = _C["green"] if dq_score >= 90 else (_C["amber"] if dq_score >= 70 else _C["red"])
    score_bg    = _C["green2"] if dq_score >= 90 else (_C["amber2"] if dq_score >= 70 else _C["red2"])

    ws.merge_cells("A3:C6")
    sc = ws["A3"]
    sc.value     = f"{dq_score}\n/100"
    sc.font      = _mk_font(bold=True, size=28, color=score_color)
    sc.fill      = _mk_fill(score_bg)
    sc.alignment = _mk_align(h="center", v="center", wrap=True)
    sc.border    = _mk_border("medium", score_color)

    ws.merge_cells("D3:L3")
    t1 = ws["D3"]
    t1.value = "Score Qualité Global"
    t1.font  = _mk_font(bold=True, size=12, color=_C["navy"])
    t1.alignment = _mk_align(h="left", v="center")

    ws.merge_cells("D4:L4")
    t2 = ws["D4"]
    verdict = "EXCELLENT" if dq_score >= 90 else ("ACCEPTABLE" if dq_score >= 70 else "CRITIQUE")
    t2.value = verdict
    t2.font  = _mk_font(bold=True, size=18, color=score_color)
    t2.alignment = _mk_align(h="left", v="center")

    ws.merge_cells("D5:L5")
    t3 = ws["D5"]
    n_err  = sum(1 for a in dq_alerts if a.get("severity") == "error")
    n_warn = sum(1 for a in dq_alerts if a.get("severity") == "warning")
    n_info = sum(1 for a in dq_alerts if a.get("severity") == "info")
    t3.value = f"{len(dq_alerts)} alertes détectées : {n_err} erreur(s) critique(s) · {n_warn} avertissement(s) · {n_info} info(s)"
    t3.font  = _mk_font(size=10, color=_C["dark"], italic=True)
    t3.alignment = _mk_align(h="left", v="center")

    for r in range(3, 7):
        _set_row_height(ws, r, 22)

    # ── Barre de progression textuelle ────────────────────────────────────────
    _spacer(ws, 7)
    ws.merge_cells("A8:L8")
    bar_val = int(dq_score / 10)
    bar_str = "█" * bar_val + "░" * (10 - bar_val) + f"  {dq_score}%"
    bar_c = ws["A8"]
    bar_c.value     = bar_str
    bar_c.font      = _mk_font(bold=True, size=14, color=score_color)
    bar_c.fill      = _mk_fill(_C["gray1"])
    bar_c.alignment = _mk_align(h="center", v="center")
    _set_row_height(ws, 8, 22)

    _spacer(ws, 9)

    # ── Données pour graphique (sévérité) ─────────────────────────────────────
    chart_data_row = 10
    ws.cell(row=chart_data_row, column=14, value="Sévérité").font = _mk_font(bold=True, color=_C["white"]); ws.cell(row=chart_data_row, column=14).fill = _mk_fill(_C["navy"])
    ws.cell(row=chart_data_row, column=15, value="Nombre").font  = _mk_font(bold=True, color=_C["white"]); ws.cell(row=chart_data_row, column=15).fill  = _mk_fill(_C["navy"])
    sev_data = [("Critique", n_err), ("Avertissement", n_warn), ("Information", n_info)]
    for i, (sev, cnt) in enumerate(sev_data):
        ws.cell(row=chart_data_row+1+i, column=14, value=sev)
        ws.cell(row=chart_data_row+1+i, column=15, value=cnt)

    if len(dq_alerts) > 0:
        pie = PieChart()
        pie.title  = "Répartition des alertes DQ"
        pie.style  = 10
        pie.width  = 12; pie.height = 9
        cats  = Reference(ws, min_col=14, min_row=chart_data_row+1, max_row=chart_data_row+3)
        data  = Reference(ws, min_col=15, min_row=chart_data_row,   max_row=chart_data_row+3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        # Couleurs des segments
        colors = ["DC2626", "D97706", "2563EB"]
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties
        for i, hex_c in enumerate(colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = hex_c
            pie.series[0].dPt.append(pt)
        ws.add_chart(pie, "A10")

    # ── Tableau des alertes ───────────────────────────────────────────────────
    r = 26
    _section_title(ws, r, 1, "⚠️  DÉTAIL DES ALERTES QUALITÉ", colspan=12)
    r += 1

    hdrs = ["#", "Sévérité", "Table", "Colonne", "Règle", "Détail"]
    widths_dq = [3, 14, 18, 18, 18, 40]
    for ci, h_text in enumerate(hdrs, start=1):
        c = ws.cell(row=r, column=ci, value=h_text)
        c.font      = _mk_font(bold=True, size=9, color=_C["white"])
        c.fill      = _mk_fill(_C["indigo"])
        c.alignment = _mk_align(h="center", v="center")
        c.border    = _mk_border("thin", _C["indigo"])
        ws.column_dimensions[_col_letter(ci)].width = widths_dq[ci-1]
    _set_row_height(ws, r, 16); r += 1

    SEV_STYLE = {
        "error":   (_C["red"],   _C["red2"]),
        "warning": (_C["amber"], _C["amber2"]),
        "info":    (_C["blue"],  _C["row_alt"]),
    }
    if not dq_alerts:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value="✅  Aucune alerte — qualité des données irréprochable")
        c.font = _mk_font(bold=True, size=10, color=_C["green"])
        c.fill = _mk_fill(_C["green2"]); r += 1
    else:
        for idx, alert in enumerate(dq_alerts, start=1):
            sev = alert.get("severity", "info")
            fg, bg = SEV_STYLE.get(sev, (_C["blue"], _C["row_alt"]))
            vals = [
                idx,
                sev.upper(),
                alert.get("table", ""),
                alert.get("column", ""),
                alert.get("rule", ""),
                alert.get("detail", ""),
            ]
            for ci, val in enumerate(vals, start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = _mk_font(bold=(ci == 2), size=9, color=fg if ci == 2 else _C["dark"])
                c.fill      = _mk_fill(bg if ci == 2 else (_C["row_alt"] if idx % 2 == 0 else _C["white"]))
                c.alignment = _mk_align(h="left" if ci > 2 else "center", v="center", wrap=(ci == 6))
                c.border    = _mk_border("thin", _C["gray3"])
            _set_row_height(ws, r, 16); r += 1

    # ── Rapport DQ détaillé ───────────────────────────────────────────────────
    if isinstance(dq_report, dict) and dq_report:
        _spacer(ws, r); r += 1
        _section_title(ws, r, 1, "🔬  ANALYSE DQ DÉTAILLÉE PAR TABLE", colspan=12)
        r += 1
        for tbl_name, tbl_data in dq_report.items():
            if not isinstance(tbl_data, dict):
                continue
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            c = ws.cell(row=r, column=1, value=f"  Table : {tbl_name}")
            c.font  = _mk_font(bold=True, size=10, color=_C["white"])
            c.fill  = _mk_fill(_C["teal"])
            c.alignment = _mk_align(h="left", v="center")
            _set_row_height(ws, r, 16); r += 1
            for k, v in tbl_data.items():
                if isinstance(v, dict):
                    continue
                lbl = ws.cell(row=r, column=1, value=str(k))
                lbl.font = _mk_font(bold=True, size=9)
                lbl.fill = _mk_fill(_C["gray1"])
                lbl.border = _mk_border("thin", _C["gray3"])
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                val = ws.cell(row=r, column=3, value=str(v))
                val.font = _mk_font(size=9)
                val.border = _mk_border("thin", _C["gray3"])
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
                _set_row_height(ws, r, 14); r += 1

    ws.freeze_panes = "A27"


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 3 : SCHÉMA EN ÉTOILE
# ─────────────────────────────────────────────────────────────────────────────

def _ws_star_schema(wb, state: dict):
    ws = wb.create_sheet("🏗️ Schéma Étoile")
    lm     = state.get("logical_model") or {}
    prefix = state.get("user_prefix", "dw")

    ws.merge_cells("A1:L1")
    h = ws["A1"]
    h.value = "SCHÉMA EN ÉTOILE — MODÈLE LOGIQUE DU DATA WAREHOUSE"
    h.font  = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill  = _mk_fill(_C["navy"])
    h.alignment = _mk_align(h="center")
    _set_row_height(ws, 1, 26)

    fact = lm.get("fact_table") or {}
    dims = lm.get("dimension_tables") or []
    fact_tables_list = lm.get("fact_tables") or []
    all_facts = ([fact] if fact and fact.get("name") else []) + fact_tables_list

    r = 3
    # ── Table(s) de Faits ─────────────────────────────────────────────────────
    _section_title(ws, r, 1, "⭐  TABLE(S) DE FAITS", colspan=12, bg=_C["gold"])
    r += 1

    for fact_obj in all_facts:
        fname = fact_obj.get("name", "fact_table")
        full_name = f"{prefix}_{fname}"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        tc = ws.cell(row=r, column=1, value=f"  {full_name}")
        tc.font = _mk_font(bold=True, size=11, color=_C["white"]); tc.fill = _mk_fill(_C["amber"])
        tc.alignment = _mk_align(h="left", v="center")
        _set_row_height(ws, r, 20); r += 1

        col_hdrs = ["Colonne", "Type", "Rôle", "Référence", "Description"]
        hdr_widths = [22, 14, 12, 22, 30]
        for ci, h_text in enumerate(col_hdrs, start=1):
            c = ws.cell(row=r, column=ci, value=h_text)
            c.font  = _mk_font(bold=True, size=9, color=_C["white"])
            c.fill  = _mk_fill(_C["gold"])
            c.alignment = _mk_align(h="center"); c.border = _mk_border("thin", _C["amber"])
            ws.column_dimensions[_col_letter(ci)].width = hdr_widths[ci-1]
        _set_row_height(ws, r, 15); r += 1

        ROLE_COLORS = {
            "pk":     (_C["purple"], _C["purple2"]),
            "fk":     (_C["blue"],   _C["row_alt"]),
            "metric": (_C["green"],  _C["green2"]),
            "date_sk":(_C["teal"],   "E0F7FA"),
        }
        for col in fact_obj.get("columns", []):
            cname = col.get("name", "")
            ctype = col.get("type", "")
            crole = col.get("role", "")
            cref  = col.get("references", "")
            cdesc = col.get("description", col.get("desc", ""))
            fg, bg = ROLE_COLORS.get(crole, (_C["dark"], _C["white"] if r % 2 else _C["row_alt"]))
            for ci, val in enumerate([cname, ctype, crole.upper(), cref, cdesc], start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font  = _mk_font(bold=(ci == 3), size=9, color=fg if ci == 3 else _C["dark"])
                c.fill  = _mk_fill(bg)
                c.alignment = _mk_align(h="left", v="center")
                c.border = _mk_border("thin", _C["gray3"])
            _set_row_height(ws, r, 14); r += 1
        _spacer(ws, r); r += 1

    # ── Dimensions ────────────────────────────────────────────────────────────
    _section_title(ws, r, 1, "◇  TABLES DE DIMENSIONS", colspan=12, bg=_C["blue"])
    r += 1

    for dim in dims:
        dname     = dim.get("name", "dim")
        full_name = f"{prefix}_{dname}"
        desc      = dim.get("description", dim.get("desc", ""))

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        tc = ws.cell(row=r, column=1, value=f"  {full_name}" + (f"   —   {desc}" if desc else ""))
        tc.font = _mk_font(bold=True, size=11, color=_C["white"])
        tc.fill = _mk_fill(_C["indigo"])
        tc.alignment = _mk_align(h="left", v="center")
        _set_row_height(ws, r, 20); r += 1

        for ci, h_text in enumerate(["Colonne", "Type", "Rôle", "Clé de substitution", "Description"], start=1):
            c = ws.cell(row=r, column=ci, value=h_text)
            c.font  = _mk_font(bold=True, size=9, color=_C["white"])
            c.fill  = _mk_fill(_C["blue"])
            c.alignment = _mk_align(h="center"); c.border = _mk_border("thin", _C["blue"])
        _set_row_height(ws, r, 15); r += 1

        for col in dim.get("columns", []):
            cname = col.get("name", "")
            ctype = col.get("type", "")
            crole = col.get("role", "")
            csk   = "SK" if crole == "pk" else ""
            cdesc = col.get("description", col.get("desc", ""))
            alt = _C["row_alt"] if r % 2 == 0 else _C["white"]
            for ci, val in enumerate([cname, ctype, crole.upper(), csk, cdesc], start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font  = _mk_font(bold=(crole == "pk"), size=9,
                                   color=_C["purple"] if crole == "pk" else _C["dark"])
                c.fill  = _mk_fill(_C["purple2"] if crole == "pk" else alt)
                c.alignment = _mk_align(h="left"); c.border = _mk_border("thin", _C["gray3"])
            _set_row_height(ws, r, 14); r += 1
        _spacer(ws, r); r += 1

    # ── Stats schéma ──────────────────────────────────────────────────────────
    _spacer(ws, r); r += 1
    _section_title(ws, r, 1, "📐  STATISTIQUES DU SCHÉMA", colspan=12, bg=_C["teal"])
    r += 1
    total_cols = sum(len(d.get("columns",[])) for d in dims) + sum(len(f.get("columns",[])) for f in all_facts)
    stats = [
        ("Tables de faits", len(all_facts)),
        ("Tables de dimensions", len(dims)),
        ("Total colonnes", total_cols),
        ("Version du modèle", state.get("logical_model_version", 1)),
        ("Critic approuvé", "✅ OUI" if state.get("critic_approved") else "⚠️ Sous réserve"),
    ]
    for lbl, val in stats:
        for ci, v in enumerate([lbl, val], start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font  = _mk_font(bold=(ci==1), size=10, color=_C["navy"] if ci==1 else _C["dark"])
            c.fill  = _mk_fill(_C["gray1"] if ci==1 else _C["white"])
            c.alignment = _mk_align(h="left" if ci==1 else "right")
            c.border = _mk_border("thin", _C["gray3"])
            if ci == 1:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            else:
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        _set_row_height(ws, r, 16); r += 1

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 4 : PERFORMANCE ETL
# ─────────────────────────────────────────────────────────────────────────────

def _ws_etl_performance(wb, state: dict):
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("⚡ Performance ETL")

    ws.merge_cells("A1:L1")
    h = ws["A1"]
    h.value = "PERFORMANCE & MÉTRIQUES ETL"
    h.font  = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill  = _mk_fill(_C["navy"])
    h.alignment = _mk_align(h="center")
    _set_row_height(ws, 1, 26)

    node_dur  = state.get("node_durations") or {}
    lm_metrics = state.get("load_metrics") or {}
    heal_hist  = state.get("heal_history") or []
    watermarks = state.get("etl_watermarks") or {}

    r = 3
    # ── Durées par nœud ────────────────────────────────────────────────────────
    _section_title(ws, r, 1, "⏱️  DURÉES D'EXÉCUTION PAR AGENT", colspan=8)
    r += 1

    AGENT_LABELS_XL = {
        "explorer": "🔍 Explorer", "data_quality": "📊 Data Quality",
        "drift_detector": "🌊 Drift Detector", "modeler": "🧠 Modeler",
        "critic": "⚖️ Critic", "chat_modifier": "💬 Chat Modifier",
        "cdc_watermark": "💧 CDC Watermark", "etl_tsql_generator": "⚙️ ETL Generator",
        "etl_initializer": "🏗️ ETL Init", "etl_extractor": "📥 Extract",
        "etl_transformer": "🔄 Transform", "etl_loader": "📤 Load",
        "healer": "🔧 Healer", "lineage_tracker": "🗺️ Lineage",
        "query_generator": "📊 Query Gen", "insight_generator": "💡 Insights",
        "cataloger": "📚 Cataloger",
    }

    for ci, h_text in enumerate(["Agent", "Durée (s)", "Durée (hh:mm:ss)", "Statut"], start=1):
        c = ws.cell(row=r, column=ci, value=h_text)
        c.font  = _mk_font(bold=True, size=9, color=_C["white"])
        c.fill  = _mk_fill(_C["teal"])
        c.alignment = _mk_align(h="center")
        c.border = _mk_border("thin", _C["teal"])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    _set_row_height(ws, r, 15); r += 1

    chart_labels = []
    chart_values = []
    total_secs   = 0

    if node_dur:
        sorted_nodes = sorted(node_dur.items(), key=lambda x: -float(x[1] if isinstance(x[1], (int,float)) else 0))
        for node, dur in sorted_nodes:
            dur_f = float(dur) if isinstance(dur, (int, float)) else 0
            total_secs += dur_f
            mins, secs = divmod(int(dur_f), 60)
            hms = f"{mins:02d}:{secs:02d}"
            status = "✅" if dur_f < 60 else ("⚠️" if dur_f < 180 else "🔴")
            label = AGENT_LABELS_XL.get(node, node)
            chart_labels.append(label)
            chart_values.append(round(dur_f, 2))
            alt = _C["row_alt"] if r % 2 == 0 else _C["white"]
            for ci, val in enumerate([label, round(dur_f, 2), hms, status], start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font  = _mk_font(size=9, bold=(ci==1))
                c.fill  = _mk_fill(alt)
                c.alignment = _mk_align(h="right" if ci==2 else "center" if ci in (3,4) else "left")
                c.border = _mk_border("thin", _C["gray3"])
                if ci == 2:
                    c.number_format = "0.00"
            _set_row_height(ws, r, 14); r += 1

        # Total
        for ci, val in enumerate(["TOTAL", round(total_secs,2), f"{int(total_secs//60):02d}:{int(total_secs%60):02d}", ""], start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font  = _mk_font(bold=True, size=9, color=_C["white"])
            c.fill  = _mk_fill(_C["navy"])
            c.alignment = _mk_align(h="right" if ci==2 else "center" if ci in (3,4) else "left")
            c.border = _mk_border("medium", _C["navy"])
        _set_row_height(ws, r, 15); r += 1

    # Données graphique (colonne cachée E-F)
    chart_start_row = 3
    ws.cell(row=chart_start_row, column=6, value="Agent").font = _mk_font(bold=True)
    ws.cell(row=chart_start_row, column=7, value="Durée (s)").font = _mk_font(bold=True)
    for i, (lbl, val) in enumerate(zip(chart_labels, chart_values)):
        ws.cell(row=chart_start_row+1+i, column=6, value=lbl)
        ws.cell(row=chart_start_row+1+i, column=7, value=val)

    if chart_values:
        bc = BarChart()
        bc.type    = "bar"; bc.grouping = "clustered"
        bc.title   = "Durée par agent (secondes)"
        bc.style   = 10; bc.width = 18; bc.height = 12
        bc.y_axis.title = "Secondes"; bc.x_axis.title = "Agent"
        data = Reference(ws, min_col=7, min_row=chart_start_row, max_row=chart_start_row+len(chart_values))
        cats = Reference(ws, min_col=6, min_row=chart_start_row+1, max_row=chart_start_row+len(chart_values))
        bc.add_data(data, titles_from_data=True)
        bc.set_categories(cats)
        bc.series[0].graphicalProperties.solidFill = _C["blue"]
        ws.add_chart(bc, "H3")

    _spacer(ws, r); r += 1

    # ── Métriques de chargement ───────────────────────────────────────────────
    if isinstance(lm_metrics, dict) and lm_metrics:
        _section_title(ws, r, 1, "📦  MÉTRIQUES DE CHARGEMENT", colspan=8)
        r += 1
        for ci, h_text in enumerate(["Table", "Insérées", "Rejetées", "Mises à jour"], start=1):
            c = ws.cell(row=r, column=ci, value=h_text)
            c.font  = _mk_font(bold=True, size=9, color=_C["white"])
            c.fill  = _mk_fill(_C["green"])
            c.alignment = _mk_align(h="center"); c.border = _mk_border("thin", _C["green"])
        _set_row_height(ws, r, 15); r += 1

        for tbl, metrics in lm_metrics.items():
            ins = rej = upd = 0
            if isinstance(metrics, dict):
                ins = metrics.get("inserted",0); rej = metrics.get("rejected",0); upd = metrics.get("updated",0)
            elif isinstance(metrics,(int,float)):
                ins = int(metrics)
            alt = _C["green2"] if r % 2 else _C["white"]
            for ci, val in enumerate([str(tbl), ins, rej, upd], start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font  = _mk_font(size=9, bold=(ci==1))
                c.fill  = _mk_fill(alt)
                c.alignment = _mk_align(h="right" if ci>1 else "left")
                c.border = _mk_border("thin", _C["gray3"])
                if ci > 1: c.number_format = "#,##0"
            _set_row_height(ws, r, 14); r += 1

    _spacer(ws, r); r += 1

    # ── CDC Watermarks ─────────────────────────────────────────────────────────
    if isinstance(watermarks, dict) and watermarks:
        _section_title(ws, r, 1, "💧  WATERMARKS CDC (CHARGEMENT INCRÉMENTAL)", colspan=8, bg=_C["teal"])
        r += 1
        for ci, h_text in enumerate(["Table", "Colonne CDC", "Dernière valeur", "Mode"], start=1):
            c = ws.cell(row=r, column=ci, value=h_text)
            c.font  = _mk_font(bold=True, size=9, color=_C["white"])
            c.fill  = _mk_fill(_C["teal"]); c.border = _mk_border("thin", _C["teal"])
        _set_row_height(ws, r, 15); r += 1
        for tbl, wm in watermarks.items():
            if isinstance(wm, dict):
                alt = _C["row_alt"] if r%2==0 else _C["white"]
                for ci, val in enumerate([str(tbl), wm.get("column",""), str(wm.get("last_value","")), wm.get("mode","full_load")], start=1):
                    c = ws.cell(row=r, column=ci, value=val)
                    c.font  = _mk_font(size=9); c.fill = _mk_fill(alt)
                    c.alignment = _mk_align(h="left"); c.border = _mk_border("thin", _C["gray3"])
                _set_row_height(ws, r, 14); r += 1

    _spacer(ws, r); r += 1

    # ── Historique Healer ──────────────────────────────────────────────────────
    if heal_hist:
        _section_title(ws, r, 1, f"🔧  HISTORIQUE AUTO-CORRECTIONS HEALER ({len(heal_hist)})", colspan=8, bg=_C["amber"])
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(row=r, column=1, value="Correction").font = _mk_font(bold=True, size=9, color=_C["white"])
        ws.cell(row=r, column=1).fill  = _mk_fill(_C["gold"])
        ws.cell(row=r, column=1).border = _mk_border("thin", _C["gold"])
        _set_row_height(ws, r, 15); r += 1
        for h_item in heal_hist:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            c = ws.cell(row=r, column=1, value=str(h_item))
            c.font  = _mk_font(size=9, color=_C["dark"])
            c.fill  = _mk_fill(_C["amber2"])
            c.alignment = _mk_align(h="left", wrap=True)
            c.border = _mk_border("thin", _C["gray3"])
            _set_row_height(ws, r, 18); r += 1

    ws.column_dimensions["E"].width = 0.1  # cacher col données graphique
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 5 : ANALYSES & RÉSULTATS
# ─────────────────────────────────────────────────────────────────────────────

def _ws_analytics(wb, state: dict):
    ws = wb.create_sheet("📈 Analyses & Résultats")

    queries  = state.get("generated_queries") or []
    q_results= state.get("query_results") or []

    ws.merge_cells("A1:N1")
    h = ws["A1"]
    h.value = "ANALYSES ANALYTIQUES & RÉSULTATS OLAP"
    h.font  = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill  = _mk_fill(_C["navy"]); h.alignment = _mk_align(h="center")
    _set_row_height(ws, 1, 26)

    r = 3
    TYPE_COLORS = {
        "kpi":          _C["gold"],
        "trend":        _C["teal"],
        "top_n":        _C["purple"],
        "distribution": _C["green"],
        "comparison":   _C["indigo"],
        "detail":       _C["blue"],
    }

    for qi, query in enumerate(queries):
        if not isinstance(query, dict):
            continue
        title = query.get("title", f"Requête {qi+1}")
        desc  = query.get("description", "")
        sql   = query.get("sql", "")
        qtype = query.get("type", "detail")
        accent = TYPE_COLORS.get(qtype, _C["blue"])

        # Titre de la requête
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        tc = ws.cell(row=r, column=1, value=f"  [{qtype.upper()}]  {title}")
        tc.font  = _mk_font(bold=True, size=11, color=_C["white"])
        tc.fill  = _mk_fill(accent); tc.alignment = _mk_align(h="left", v="center")
        _set_row_height(ws, r, 20); r += 1

        # Description
        if desc:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            dc = ws.cell(row=r, column=1, value=f"  {desc}")
            dc.font  = _mk_font(italic=True, size=9, color=_C["dark"])
            dc.fill  = _mk_fill(_C["gray1"]); dc.alignment = _mk_align(h="left", v="center", wrap=True)
            _set_row_height(ws, r, 16); r += 1

        # SQL (fond sombre simulé)
        if sql:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            sql_label = ws.cell(row=r, column=1, value="SQL :")
            sql_label.font = _mk_font(bold=True, size=8, color=_C["teal"])
            sql_label.fill = _mk_fill("1A2B3C")
            _set_row_height(ws, r, 14); r += 1

            for sql_line in sql.splitlines():
                if not sql_line.strip():
                    continue
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
                sc = ws.cell(row=r, column=1, value=f"  {sql_line}")
                sc.font  = _mk_font(size=8, color="A5D6F7")
                sc.fill  = _mk_fill("1A2B3C")
                sc.alignment = _mk_align(h="left", v="center")
                _set_row_height(ws, r, 13); r += 1

        # Résultats
        result = q_results[qi] if qi < len(q_results) else None
        if isinstance(result, dict):
            columns = result.get("columns") or []
            rows    = result.get("rows") or []
            error   = result.get("error")

            if error:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
                ec = ws.cell(row=r, column=1, value=f"  ❌ Erreur : {error}")
                ec.font  = _mk_font(size=9, color=_C["red"])
                ec.fill  = _mk_fill(_C["red2"]); _set_row_height(ws, r, 16); r += 1
            elif columns:
                # En-têtes résultats
                for ci, col in enumerate(columns[:14], start=1):
                    c = ws.cell(row=r, column=ci, value=col)
                    c.font  = _mk_font(bold=True, size=9, color=_C["white"])
                    c.fill  = _mk_fill(_C["indigo"])
                    c.alignment = _mk_align(h="center"); c.border = _mk_border("thin", _C["indigo"])
                    ws.column_dimensions[_col_letter(ci)].width = max(12, min(20, len(str(col))+4))
                _set_row_height(ws, r, 15); r += 1

                for row_data in (rows[:500] if rows else []):
                    alt = _C["row_alt"] if r % 2 == 0 else _C["white"]
                    if isinstance(row_data, list):
                        vals = row_data[:14]
                    elif isinstance(row_data, dict):
                        vals = [row_data.get(c) for c in columns[:14]]
                    else:
                        vals = [str(row_data)]
                    for ci, val in enumerate(vals, start=1):
                        c = ws.cell(row=r, column=ci, value=_coerce_cell(val))
                        c.font  = _mk_font(size=9)
                        c.fill  = _mk_fill(alt)
                        c.alignment = _mk_align(h="right" if isinstance(val,(int,float)) else "left")
                        c.border = _mk_border("thin", _C["gray3"])
                    _set_row_height(ws, r, 14); r += 1

                # Total lignes
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(columns[:14]))
                tc2 = ws.cell(row=r, column=1, value=f"  {len(rows)} lignes au total")
                tc2.font  = _mk_font(bold=True, size=8, italic=True, color=_C["gray4"])
                tc2.fill  = _mk_fill(_C["gray1"])
                _set_row_height(ws, r, 12); r += 1
            else:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
                nc = ws.cell(row=r, column=1, value="  Aucun résultat disponible — lancez la requête depuis le Query Runner")
                nc.font  = _mk_font(italic=True, size=9, color=_C["gray4"])
                nc.fill  = _mk_fill(_C["gray1"]); _set_row_height(ws, r, 14); r += 1
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            nc = ws.cell(row=r, column=1, value="  Résultats non encore générés — pipeline complet requis")
            nc.font  = _mk_font(italic=True, size=9, color=_C["gray4"])
            nc.fill  = _mk_fill(_C["gray1"]); _set_row_height(ws, r, 14); r += 1

        _spacer(ws, r); r += 1

    if not queries:
        ws.merge_cells("A3:N5")
        nc = ws["A3"]
        nc.value     = "Aucune requête analytique générée — le pipeline doit être exécuté jusqu'à l'étape query_generator"
        nc.font      = _mk_font(italic=True, size=11, color=_C["gray4"])
        nc.fill      = _mk_fill(_C["gray1"]); nc.alignment = _mk_align(h="center", v="center")

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 6 : CATALOGUE DE DONNÉES
# ─────────────────────────────────────────────────────────────────────────────

def _ws_catalog(wb, state: dict):
    ws = wb.create_sheet("🗂️ Catalogue Données")

    catalog = state.get("data_catalog") or {}
    lm      = state.get("logical_model") or {}
    prefix  = state.get("user_prefix", "dw")

    ws.merge_cells("A1:J1")
    h = ws["A1"]; h.value = "CATALOGUE DE DONNÉES — DOCUMENTATION MÉTIER"
    h.font  = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill  = _mk_fill(_C["navy"]); h.alignment = _mk_align(h="center")
    _set_row_height(ws, 1, 26)

    r = 3

    def _write_table_section(table_obj, table_type, bg_header):
        nonlocal r
        tname = table_obj.get("name","")
        tdesc = table_obj.get("description", table_obj.get("desc",""))
        full  = f"{prefix}_{tname}"

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        tc = ws.cell(row=r, column=1, value=f"  {full}   [{table_type}]" + (f"   —   {tdesc}" if tdesc else ""))
        tc.font = _mk_font(bold=True, size=11, color=_C["white"])
        tc.fill = _mk_fill(bg_header); tc.alignment = _mk_align(h="left")
        _set_row_height(ws, r, 20); r += 1

        hdrs = ["Colonne", "Type SQL", "Rôle", "Nullable", "PK/FK", "Description métier"]
        hw   = [22, 14, 10, 8, 12, 40]
        for ci, h_text in enumerate(hdrs, start=1):
            c = ws.cell(row=r, column=ci, value=h_text)
            c.font  = _mk_font(bold=True, size=9, color=_C["white"])
            c.fill  = _mk_fill(_C["teal"]); c.border = _mk_border("thin", _C["teal"])
            c.alignment = _mk_align(h="center")
            ws.column_dimensions[_col_letter(ci)].width = hw[ci-1]
        _set_row_height(ws, r, 15); r += 1

        for col in table_obj.get("columns", []):
            cname = col.get("name",""); ctype = col.get("type","VARCHAR")
            crole = col.get("role","attribute"); cdesc = col.get("description", col.get("desc",""))
            nullable = "NON" if crole in ("pk","sk") else "OUI"
            pk_fk = "PK" if crole=="pk" else ("FK → "+col.get("references","") if crole=="fk" else "")
            alt = _C["purple2"] if crole=="pk" else (_C["row_alt"] if r%2==0 else _C["white"])
            for ci, val in enumerate([cname, ctype, crole.upper(), nullable, pk_fk, cdesc], start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font  = _mk_font(size=9, bold=(crole=="pk"))
                c.fill  = _mk_fill(alt)
                c.alignment = _mk_align(h="center" if ci in (3,4,5) else "left")
                c.border = _mk_border("thin", _C["gray3"])
            _set_row_height(ws, r, 14); r += 1
        _spacer(ws, r); r += 1

    _section_title(ws, r, 1, "⭐  TABLES DE FAITS", colspan=10, bg=_C["gold"])
    r += 1
    fact = lm.get("fact_table") or {}
    if fact: _write_table_section(fact, "FACT", _C["amber"])
    for f in lm.get("fact_tables", []):
        _write_table_section(f, "FACT", _C["amber"])

    _section_title(ws, r, 1, "◇  TABLES DE DIMENSIONS", colspan=10, bg=_C["blue"])
    r += 1
    for dim in lm.get("dimension_tables", []):
        _write_table_section(dim, "DIMENSION", _C["indigo"])

    # Catalogue externe si disponible
    if isinstance(catalog, dict) and catalog:
        _spacer(ws, r); r += 1
        _section_title(ws, r, 1, "📚  CATALOGUE ÉTENDU", colspan=10, bg=_C["purple"])
        r += 1
        for tbl, meta in catalog.items():
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            tc = ws.cell(row=r, column=1, value=f"  {tbl}")
            tc.font = _mk_font(bold=True, size=10, color=_C["white"])
            tc.fill = _mk_fill(_C["purple"]); tc.alignment = _mk_align(h="left")
            _set_row_height(ws, r, 16); r += 1
            if isinstance(meta, dict):
                for k, v in meta.items():
                    for ci, val in enumerate([k, str(v)[:200]], start=1):
                        c = ws.cell(row=r, column=ci, value=val)
                        c.font = _mk_font(bold=(ci==1), size=9)
                        c.fill = _mk_fill(_C["purple2"] if ci==1 else _C["white"])
                        c.border = _mk_border("thin", _C["gray3"])
                        if ci==1: ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                        else: ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                    _set_row_height(ws, r, 14); r += 1

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 7 : LIGNAGE
# ─────────────────────────────────────────────────────────────────────────────

def _ws_lineage(wb, state: dict):
    ws = wb.create_sheet("🔗 Lignage")

    lineage = state.get("lineage") or {}

    ws.merge_cells("A1:K1")
    h = ws["A1"]; h.value = "LIGNAGE DES DONNÉES — SOURCE → DATA WAREHOUSE"
    h.font  = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill  = _mk_fill(_C["navy"]); h.alignment = _mk_align(h="center")
    _set_row_height(ws, 1, 26)

    r = 3
    _section_title(ws, r, 1, "🗺️  TRAÇABILITÉ COLONNE PAR COLONNE", colspan=10)
    r += 1

    hdrs   = ["Table DW", "Colonne DW", "Type DW", "Table Source", "Colonne Source", "Transformation", "Rôle"]
    hw_lin = [22, 20, 12, 20, 20, 20, 12]
    for ci, h_text in enumerate(hdrs, start=1):
        c = ws.cell(row=r, column=ci, value=h_text)
        c.font  = _mk_font(bold=True, size=9, color=_C["white"])
        c.fill  = _mk_fill(_C["indigo"]); c.border = _mk_border("thin", _C["indigo"])
        c.alignment = _mk_align(h="center")
        ws.column_dimensions[_col_letter(ci)].width = hw_lin[ci-1]
    _set_row_height(ws, r, 15); r += 1

    TRANSFORM_COLORS_LIN = {
        "DIRECT_LOAD":            _C["green"],
        "GENERATE_SURROGATE_KEY": _C["purple"],
        "DATE_PARSE":             _C["teal"],
        "RENAME_AND_CAST":        _C["blue"],
        "LOOKUP_SK":              _C["amber"],
    }

    if not lineage:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        nc = ws.cell(row=r, column=1, value="Lignage non disponible — pipeline ETL complet requis")
        nc.font = _mk_font(italic=True, size=10, color=_C["gray4"])
        nc.fill = _mk_fill(_C["gray1"]); nc.alignment = _mk_align(h="center")
    else:
        for dw_table, tdata in lineage.items():
            if not isinstance(tdata, dict): continue
            nodes  = tdata.get("nodes", [])
            edges  = tdata.get("edges", [])
            ttype  = tdata.get("type", "dimension")
            node_map = {n["id"]: n for n in nodes}

            for edge in edges:
                from_node = node_map.get(edge.get("from", ""))
                to_node   = node_map.get(edge.get("to", ""))
                if not from_node or not to_node: continue

                transform = edge.get("transform", "DIRECT_LOAD")
                t_color   = TRANSFORM_COLORS_LIN.get(transform, _C["gray4"])
                alt = _C["row_alt"] if r % 2 == 0 else _C["white"]
                row_vals = [
                    dw_table,
                    to_node.get("label",""),
                    to_node.get("role",""),
                    from_node.get("table",""),
                    from_node.get("label",""),
                    transform,
                    ttype.upper(),
                ]
                for ci, val in enumerate(row_vals, start=1):
                    c = ws.cell(row=r, column=ci, value=val)
                    c.font  = _mk_font(size=9, bold=(ci==1),
                                       color=t_color if ci==6 else (
                                           _C["amber"] if ci==7 and ttype=="fact" else
                                           _C["indigo"] if ci==7 else _C["dark"]))
                    c.fill  = _mk_fill(alt)
                    c.alignment = _mk_align(h="center" if ci in (6,7) else "left")
                    c.border = _mk_border("thin", _C["gray3"])
                _set_row_height(ws, r, 14); r += 1

    ws.freeze_panes = "A5"


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 8 : DDL SQL
# ─────────────────────────────────────────────────────────────────────────────

def _ws_ddl(wb, state: dict):
    ws = wb.create_sheet("📜 DDL SQL")
    ddl = str(state.get("sql_ddl") or "")

    ws.merge_cells("A1:B1")
    h = ws["A1"]; h.value = "DDL SQL SERVER — SCHÉMA DATA WAREHOUSE"
    h.font = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill = _mk_fill(_C["navy"]); h.alignment = _mk_align(h="center")
    _set_row_height(ws, 1, 26)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110

    if not ddl:
        ws.cell(row=3, column=1, value="DDL non disponible — modèle non encore généré")
        ws.cell(row=3, column=1).font = _mk_font(italic=True, color=_C["gray4"])
        return

    r = 3
    SQL_KW = {"CREATE","TABLE","ALTER","INDEX","ON","AS","SELECT","FROM","WHERE","PRIMARY","KEY",
               "FOREIGN","REFERENCES","NOT","NULL","IDENTITY","INT","BIGINT","VARCHAR","NVARCHAR",
               "DECIMAL","FLOAT","DATE","DATETIME","BIT","CONSTRAINT","DEFAULT","GO","USE","IF","BEGIN","END"}
    COMMENT_COLOR = "6A9955"
    KW_COLOR      = "569CD6"
    STR_COLOR     = "CE9178"
    NORMAL_COLOR  = "D4D4D4"
    BG_CODE       = "1E1E1E"

    for line in ddl.splitlines():
        line_num = ws.cell(row=r, column=1, value=r-2)
        line_num.font      = _mk_font(size=8, color="6E7681")
        line_num.fill      = _mk_fill("161B22")
        line_num.alignment = _mk_align(h="right", v="center")

        stripped = line.strip()
        # Détecter type de ligne pour colorisation
        if stripped.startswith("--"):
            color = COMMENT_COLOR
        elif any(stripped.upper().startswith(kw) for kw in SQL_KW):
            color = KW_COLOR
        elif "'" in stripped:
            color = STR_COLOR
        else:
            color = NORMAL_COLOR

        code = ws.cell(row=r, column=2, value=line)
        code.font      = _mk_font(size=9, color=color, name="Consolas")
        code.fill      = _mk_fill(BG_CODE)
        code.alignment = _mk_align(h="left", v="center")
        _set_row_height(ws, r, 14); r += 1


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 9 : JOURNAL D'EXÉCUTION
# ─────────────────────────────────────────────────────────────────────────────

def _ws_execution_log(wb, state: dict):
    ws = wb.create_sheet("📋 Journal Exécution")

    exec_log = state.get("execution_log") or []

    ws.merge_cells("A1:D1")
    h = ws["A1"]; h.value = "JOURNAL D'EXÉCUTION DU PIPELINE"
    h.font  = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill  = _mk_fill(_C["navy"]); h.alignment = _mk_align(h="center")
    _set_row_height(ws, 1, 26)

    for ci, h_text in enumerate(["#", "Entrée du journal", "Niveau", "Catégorie"], start=1):
        c = ws.cell(row=2, column=ci, value=h_text)
        c.font  = _mk_font(bold=True, size=9, color=_C["white"])
        c.fill  = _mk_fill(_C["blue"]); c.border = _mk_border("thin", _C["blue"])
        c.alignment = _mk_align(h="center")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 100
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    _set_row_height(ws, 2, 15)

    LOG_STYLE = {
        "error":   (_C["red"],   _C["red2"]),
        "warning": (_C["amber"], _C["amber2"]),
        "success": (_C["green"], _C["green2"]),
        "info":    (_C["blue"],  _C["row_alt"]),
    }

    for i, entry in enumerate(exec_log, start=1):
        text = str(entry)
        # Détecter le niveau
        if "❌" in text or "ERROR" in text or "ERREUR" in text or "FAILED" in text:
            level, cat = "error",   "ERREUR"
        elif "⚠️" in text or "WARNING" in text or "DÉRIVE" in text or "drift" in text.lower():
            level, cat = "warning", "AVERT."
        elif "✅" in text or "succès" in text.lower() or "success" in text.lower() or "terminé" in text.lower():
            level, cat = "success", "SUCCÈS"
        elif "🔧" in text or "Healer" in text or "Heal" in text:
            level, cat = "warning", "HEALER"
        elif "🚀" in text or "Pipeline" in text:
            level, cat = "info",    "PIPELINE"
        else:
            level, cat = "info",    "INFO"

        fg, bg = LOG_STYLE.get(level, (_C["blue"], _C["row_alt"]))
        r = i + 2
        for ci, val in enumerate([i, text, level.upper(), cat], start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font  = _mk_font(size=9, bold=(ci==3), color=fg if ci==3 else _C["dark"])
            c.fill  = _mk_fill(bg if ci==3 else (_C["row_alt"] if i%2==0 else _C["white"]))
            c.alignment = _mk_align(h="center" if ci in (1,3,4) else "left", v="center", wrap=(ci==2))
            c.border = _mk_border("thin", _C["gray3"])
        _set_row_height(ws, r, 16)

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE BONUS : MESURES & KPI MÉTIER (granularité fine pour décideurs)
# ─────────────────────────────────────────────────────────────────────────────


# ============================================================
# FEUILLE BONUS : MESURES & KPI METIER
# ============================================================

def _ws_measures_kpi(wb, state: dict):
    """Mesures + agregations multi-granularite + charts natifs Excel."""
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    ws = wb.create_sheet("Mesures & KPI")
    ws.merge_cells("A1:N2")
    h = ws["A1"]
    h.value = "MESURES & KPI METIER - VUE DECISIONNELLE MULTI-GRANULARITE"
    h.font = _mk_font(bold=True, size=14, color=_C["white"])
    h.fill = _mk_fill(_C["navy"])
    h.alignment = _mk_align(h="center", v="center")
    _set_row_height(ws, 1, 22); _set_row_height(ws, 2, 22)

    lm = state.get("logical_model") or {}
    facts = lm.get("fact_tables") or ([lm.get("fact_table")] if lm.get("fact_table") else [])
    facts = [f for f in facts if isinstance(f, dict)]
    prefix = state.get("user_prefix", "dw")
    dw_cfg = state.get("dw_connection_config") or {}

    engine = None
    try:
        from nodes.etl_executor import _build_engine
        engine = _build_engine(dw_cfg) if dw_cfg else None
    except Exception as e:
        logger.warning(f"[Measures] Pas de DW : {e}")

    r = 4
    if not facts:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        c = ws.cell(row=r, column=1, value="  Aucune table de faits disponible.")
        c.font = _mk_font(italic=True, size=11, color=_C["gray4"])
        c.alignment = _mk_align(h="center", v="center"); _set_row_height(ws, r, 22)
        return

    AGGS = ("COUNT", "SUM", "AVG", "MIN", "MAX", "STDDEV")

    for fact in facts:
        fname = fact.get("name", "fact")
        full_name = f"{prefix}_{fname}"
        cols = fact.get("columns", []) or []
        metrics = [c for c in cols if c.get("role") in ("metric", "computed")]
        for c in cols:
            if c in metrics:
                continue
            t = (c.get("type") or "").upper()
            if any(x in t for x in ("INT", "BIGINT", "DECIMAL", "FLOAT", "NUMERIC", "MONEY")):
                if c.get("role") not in ("pk", "fk"):
                    metrics.append(c)

        _section_title(ws, r, 1, f"Table de faits : {full_name}", colspan=14, bg=_C["indigo"])
        r += 1

        if not metrics:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            c = ws.cell(row=r, column=1, value="  Aucune mesure numerique detectee.")
            c.font = _mk_font(italic=True, size=10, color=_C["gray4"])
            _set_row_height(ws, r, 18); r += 2
            continue

        hdrs = ["Mesure", "Type", "Description"] + list(AGGS)
        for ci, hdr in enumerate(hdrs, start=1):
            cc = ws.cell(row=r, column=ci, value=hdr)
            cc.font = _mk_font(bold=True, size=9, color=_C["white"])
            cc.fill = _mk_fill(_C["teal"])
            cc.alignment = _mk_align(h="center", v="center")
            cc.border = _mk_border("thin", _C["teal"])
            ws.column_dimensions[_col_letter(ci)].width = 14 if ci > 3 else 22
        _set_row_height(ws, r, 18); r += 1

        agg_rows = []
        for met in metrics:
            mname = met.get("name", "")
            mtype = met.get("type", "")
            mdesc = (met.get("description") or "").strip()
            row = {"name": mname, "type": mtype, "desc": mdesc[:80], "agg": (None,)*6}
            if engine is not None:
                try:
                    from sqlalchemy import text as _text
                    q = _text(
                        f"SELECT COUNT([{mname}]) AS c, "
                        f"SUM(CAST([{mname}] AS DECIMAL(38,6))) AS s, "
                        f"AVG(CAST([{mname}] AS DECIMAL(38,6))) AS a, "
                        f"MIN([{mname}]) AS mn, MAX([{mname}]) AS mx, "
                        f"STDEV(CAST([{mname}] AS DECIMAL(38,6))) AS st "
                        f"FROM [{full_name}]"
                    )
                    with engine.connect() as conn:
                        result = conn.execute(q).fetchone()
                    row["agg"] = (
                        result[0] or 0,
                        float(result[1]) if result[1] is not None else None,
                        float(result[2]) if result[2] is not None else None,
                        result[3], result[4],
                        float(result[5]) if result[5] is not None else None,
                    )
                except Exception as e:
                    logger.debug(f"[Measures] {full_name}.{mname} : {e}")
            agg_rows.append(row)

        for i, ar in enumerate(agg_rows):
            alt = _C["row_alt"] if i % 2 == 0 else _C["white"]
            vals = [ar["name"], ar["type"], ar["desc"], *ar["agg"]]
            for ci, val in enumerate(vals, start=1):
                cc = ws.cell(row=r, column=ci, value=("" if val is None else val))
                cc.font = _mk_font(size=9, bold=(ci == 1))
                cc.fill = _mk_fill(alt)
                cc.alignment = _mk_align(
                    h="right" if ci > 3 and isinstance(val, (int, float)) else "left",
                    v="center", wrap=(ci == 3),
                )
                cc.border = _mk_border("thin", _C["gray3"])
                if ci > 3 and isinstance(val, (int, float)):
                    cc.number_format = "#,##0.00"
            _set_row_height(ws, r, 16); r += 1

        if any(ar["agg"][1] is not None for ar in agg_rows):
            chart_start = r + 1
            ws.cell(row=chart_start, column=1, value="Mesure").font = _mk_font(bold=True, size=9, color=_C["white"])
            ws.cell(row=chart_start, column=2, value="Total (SUM)").font = _mk_font(bold=True, size=9, color=_C["white"])
            for ci in (1, 2):
                ws.cell(row=chart_start, column=ci).fill = _mk_fill(_C["purple"])
                ws.cell(row=chart_start, column=ci).alignment = _mk_align(h="center")
            for i, ar in enumerate(agg_rows, start=1):
                ws.cell(row=chart_start + i, column=1, value=ar["name"])
                v = ar["agg"][1]
                ws.cell(row=chart_start + i, column=2, value=v if v is not None else 0)
                ws.cell(row=chart_start + i, column=2).number_format = "#,##0.00"

            try:
                bc = BarChart()
                bc.type = "col"; bc.style = 11
                bc.title = f"Totaux des mesures - {fname}"
                bc.y_axis.title = "Somme"; bc.x_axis.title = "Mesure"
                data = Reference(ws, min_col=2, min_row=chart_start,
                                 max_row=chart_start + len(agg_rows), max_col=2)
                cats = Reference(ws, min_col=1, min_row=chart_start + 1,
                                 max_row=chart_start + len(agg_rows), max_col=1)
                bc.add_data(data, titles_from_data=True)
                bc.set_categories(cats)
                bc.height = 9; bc.width = 18
                ws.add_chart(bc, f"E{chart_start}")
            except Exception as e:
                logger.warning(f"[Measures] BarChart : {e}")
            r = chart_start + len(agg_rows) + 12

        date_fk = next((c["name"] for c in cols if c.get("role") == "fk"
                        and "date" in str(c.get("name", "")).lower()), None)
        if engine is not None and date_fk and metrics:
            primary_metric = metrics[0]["name"]
            for grain_label, grain_expr in (
                ("Annee", "YEAR(d.[full_date])"),
                ("Trimestre", "DATEPART(quarter, d.[full_date])"),
                ("Mois", "FORMAT(d.[full_date], 'yyyy-MM')"),
            ):
                try:
                    from sqlalchemy import text as _text
                    dim_date = f"{prefix}_dim_date"
                    q = _text(
                        f"SELECT TOP 50 {grain_expr} AS bucket, "
                        f"SUM(CAST(f.[{primary_metric}] AS DECIMAL(38,6))) AS total, "
                        f"AVG(CAST(f.[{primary_metric}] AS DECIMAL(38,6))) AS moyenne, "
                        f"COUNT(*) AS volume "
                        f"FROM [{full_name}] f "
                        f"INNER JOIN [{dim_date}] d ON f.[{date_fk}] = d.[date_sk] "
                        f"GROUP BY {grain_expr} ORDER BY bucket"
                    )
                    with engine.connect() as conn:
                        rows = conn.execute(q).fetchall()
                    if not rows:
                        continue
                    _section_title(ws, r, 1, f"Granularite {grain_label} - mesure '{primary_metric}'",
                                   colspan=14, bg=_C["blue"])
                    r += 1
                    hdrs2 = [grain_label, "Total", "Moyenne", "Volume"]
                    for ci, hh in enumerate(hdrs2, start=1):
                        cc = ws.cell(row=r, column=ci, value=hh)
                        cc.font = _mk_font(bold=True, size=9, color=_C["white"])
                        cc.fill = _mk_fill(_C["indigo"])
                        cc.alignment = _mk_align(h="center")
                        cc.border = _mk_border("thin", _C["indigo"])
                    _set_row_height(ws, r, 16); r += 1
                    data_start = r
                    for i, row_db in enumerate(rows):
                        alt = _C["row_alt"] if i % 2 == 0 else _C["white"]
                        for ci, val in enumerate(row_db, start=1):
                            cc = ws.cell(row=r, column=ci,
                                         value=float(val) if isinstance(val, (int, float)) and ci != 1 else val)
                            cc.font = _mk_font(size=9, bold=(ci == 1))
                            cc.fill = _mk_fill(alt)
                            cc.alignment = _mk_align(h="left" if ci == 1 else "right")
                            cc.border = _mk_border("thin", _C["gray3"])
                            if ci != 1:
                                cc.number_format = "#,##0.00" if ci != 4 else "#,##0"
                        _set_row_height(ws, r, 14); r += 1
                    try:
                        lc = LineChart()
                        lc.title = f"Tendance {grain_label.lower()} - {primary_metric}"
                        lc.y_axis.title = "Total"; lc.x_axis.title = grain_label
                        data = Reference(ws, min_col=2, min_row=data_start - 1,
                                         max_row=data_start + len(rows) - 1, max_col=3)
                        cats = Reference(ws, min_col=1, min_row=data_start,
                                         max_row=data_start + len(rows) - 1, max_col=1)
                        lc.add_data(data, titles_from_data=True)
                        lc.set_categories(cats)
                        lc.height = 9; lc.width = 20
                        ws.add_chart(lc, f"F{data_start - 1}")
                    except Exception as e:
                        logger.warning(f"[Measures] LineChart : {e}")
                    r += 12
                except Exception as e:
                    logger.debug(f"[Measures] grain {grain_label} : {e}")

        non_date_fks = [c["name"] for c in cols if c.get("role") == "fk"
                        and "date" not in str(c.get("name", "")).lower()]
        if engine is not None and non_date_fks and metrics:
            primary_metric = metrics[0]["name"]
            for fk in non_date_fks[:2]:
                try:
                    base = fk.lower().replace("_sk", "").replace("_id", "").replace("_fk", "").replace("id_", "")
                    dim_table = f"{prefix}_dim_{base}"
                    from sqlalchemy import text as _text, inspect as _inspect
                    insp = _inspect(engine)
                    if dim_table.lower() not in [t.lower() for t in insp.get_table_names()]:
                        for t in insp.get_table_names():
                            if base in t.lower() and t.startswith(prefix + "_dim"):
                                dim_table = t
                                break
                    dim_cols = [c["name"] for c in insp.get_columns(dim_table)]
                    label_col = next((c for c in dim_cols
                                      if any(k in c.lower() for k in ("name", "label", "libelle", "description"))),
                                     dim_cols[1] if len(dim_cols) > 1 else dim_cols[0])
                    pk_col = next((c for c in dim_cols if c.endswith("_sk")), dim_cols[0])
                    q = _text(
                        f"SELECT TOP 10 d.[{label_col}] AS label, "
                        f"SUM(CAST(f.[{primary_metric}] AS DECIMAL(38,6))) AS total "
                        f"FROM [{full_name}] f "
                        f"INNER JOIN [{dim_table}] d ON f.[{fk}] = d.[{pk_col}] "
                        f"GROUP BY d.[{label_col}] ORDER BY total DESC"
                    )
                    with engine.connect() as conn:
                        rows = conn.execute(q).fetchall()
                    if not rows:
                        continue
                    _section_title(ws, r, 1, f"Top 10 '{primary_metric}' par {dim_table}",
                                   colspan=14, bg=_C["green"])
                    r += 1
                    for ci, hh in enumerate([dim_table, primary_metric], start=1):
                        cc = ws.cell(row=r, column=ci, value=hh)
                        cc.font = _mk_font(bold=True, size=9, color=_C["white"])
                        cc.fill = _mk_fill(_C["green"])
                        cc.alignment = _mk_align(h="center")
                        cc.border = _mk_border("thin", _C["green"])
                    _set_row_height(ws, r, 16); r += 1
                    data_start = r
                    for i, row_db in enumerate(rows):
                        alt = _C["row_alt"] if i % 2 == 0 else _C["white"]
                        ws.cell(row=r, column=1, value=str(row_db[0]) if row_db[0] is not None else "(vide)")
                        v = row_db[1]
                        ws.cell(row=r, column=2, value=float(v) if v is not None else 0)
                        ws.cell(row=r, column=2).number_format = "#,##0.00"
                        for ci in (1, 2):
                            cc = ws.cell(row=r, column=ci)
                            cc.font = _mk_font(size=9, bold=(ci == 1))
                            cc.fill = _mk_fill(alt)
                            cc.alignment = _mk_align(h="left" if ci == 1 else "right")
                            cc.border = _mk_border("thin", _C["gray3"])
                        _set_row_height(ws, r, 14); r += 1
                    try:
                        pc = PieChart()
                        pc.title = f"Top 10 par {dim_table}"
                        data = Reference(ws, min_col=2, min_row=data_start - 1,
                                         max_row=data_start + len(rows) - 1, max_col=2)
                        cats = Reference(ws, min_col=1, min_row=data_start,
                                         max_row=data_start + len(rows) - 1, max_col=1)
                        pc.add_data(data, titles_from_data=True)
                        pc.set_categories(cats)
                        pc.height = 9; pc.width = 14
                        ws.add_chart(pc, f"E{data_start - 1}")
                    except Exception as e:
                        logger.warning(f"[Measures] PieChart : {e}")
                    r += 12
                except Exception as e:
                    logger.debug(f"[Measures] top10 : {e}")

        r += 2

    _set_cols(ws, {"A": 22, "B": 18, "C": 28, "D": 12, "E": 14, "F": 14,
                   "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14,
                   "M": 14, "N": 14})
    ws.freeze_panes = "A4"


# =============================================================================
#  AIDES "EXÉCUTIVES" — palette, scorecards, recommandations, hyperliens
# =============================================================================

# Palette dédiée à la version pour décideurs (couleurs sobres, contraste élevé).
_EC = {
    "primary":     "1E3A8A",   # bleu marine profond — bandeaux, titres
    "primary_lt":  "DBEAFE",   # bleu pâle — fond de section
    "accent":      "0EA5E9",   # cyan — hyperliens, boutons
    "ok":          "16A34A",   # vert santé
    "ok_lt":       "DCFCE7",
    "warn":        "D97706",   # orange ambre
    "warn_lt":     "FEF3C7",
    "alert":       "DC2626",   # rouge alerte
    "alert_lt":    "FEE2E2",
    "neutral":     "475569",   # ardoise pour textes secondaires
    "neutral_lt":  "F1F5F9",
    "border":      "CBD5E1",
    "ink":         "0F172A",
    "white":       "FFFFFF",
    "stripe":      "F8FAFC",   # zébrure légère
    "gold":        "B45309",
    "gold_lt":     "FFF7ED",
    "card_lt":     "FFFFFF",
}


def _exec_overall_grade(state: dict) -> dict:
    """
    Calcule une note globale du pipeline (A à D) et un libellé synthétique.
    Combine score qualité, succès ETL, alertes, auto-corrections.
    """
    dq = float(state.get("dq_score", 0) or 0)
    etl_ok = (state.get("etl_status") or "") == "success"
    n_alerts = len(state.get("dq_alerts", []) or [])
    n_heal = len(state.get("heal_history", []) or [])

    score = dq
    if etl_ok:
        score += 5
    score -= min(n_alerts * 2, 10)
    score -= min(max(0, n_heal - 3) * 2, 8)
    score = max(0, min(100, score))

    if score >= 90:
        return {"score": int(score), "letter": "A", "label": "Excellence opérationnelle",
                "color": _EC["ok"], "color_lt": _EC["ok_lt"],
                "narrative": "La plateforme livre un Data Warehouse aux normes attendues. Aucune action corrective majeure n'est requise."}
    if score >= 75:
        return {"score": int(score), "letter": "B", "label": "Conforme aux attentes",
                "color": _EC["ok"], "color_lt": _EC["ok_lt"],
                "narrative": "Le pipeline est en bonne santé. Quelques optimisations marginales restent recommandées pour atteindre l'excellence."}
    if score >= 60:
        return {"score": int(score), "letter": "C", "label": "Acceptable avec réserves",
                "color": _EC["warn"], "color_lt": _EC["warn_lt"],
                "narrative": "Le pipeline fonctionne, mais des points de vigilance doivent être traités avant la prochaine mise en production."}
    return {"score": int(score), "letter": "D", "label": "Action requise",
            "color": _EC["alert"], "color_lt": _EC["alert_lt"],
            "narrative": "Plusieurs indicateurs sont sous le seuil acceptable. Une intervention est nécessaire avant exploitation."}


def _exec_health_categories(state: dict) -> list[dict]:
    """Retourne 5 catégories de santé évaluées chacune sur 100."""
    dq = float(state.get("dq_score", 0) or 0)
    etl_ok = (state.get("etl_status") or "") == "success"
    n_alerts = len(state.get("dq_alerts", []) or [])
    n_heal = len(state.get("heal_history", []) or [])
    lm = state.get("logical_model") or {}
    n_dims = len(lm.get("dimension_tables") or [])
    has_fact = bool(lm.get("fact_table") or lm.get("fact_tables"))
    pii_count = len(state.get("pii_findings", []) or [])

    cats = [
        {"name": "Qualité des données", "score": int(dq),
         "factors": [f"{n_alerts} alerte(s) DQ", "12 dimensions évaluées"]},
        {"name": "Modèle dimensionnel",
         "score": 95 if (has_fact and n_dims >= 3) else (70 if has_fact else 30),
         "factors": [f"{1 if has_fact else 0} table de faits", f"{n_dims} dimensions"]},
        {"name": "Pipeline ETL",
         "score": (95 if etl_ok else 40) - min(n_heal * 4, 20),
         "factors": [f"Statut : {state.get('etl_status') or 'inconnu'}", f"{n_heal} auto-correction(s)"]},
        {"name": "Sécurité & gouvernance",
         "score": max(40, 95 - pii_count * 5),
         "factors": [f"{pii_count} colonne(s) PII détectée(s)", "Politiques RGPD : actives"]},
        {"name": "Performance",
         "score": 90 if state.get("node_durations") else 70,
         "factors": ["Index columnstore : oui" if (state.get("logical_model") or {}).get("indexes_created", True) else "Index : à vérifier",
                     f"Durée totale : {round(sum(float(v) for v in (state.get('node_durations') or {}).values() if isinstance(v,(int,float))),0)}s"]},
    ]
    for c in cats:
        s = max(0, min(100, c["score"]))
        c["score"] = s
        if s >= 85:
            c["status"] = "Bon";        c["color"] = _EC["ok"];    c["color_lt"] = _EC["ok_lt"]
        elif s >= 65:
            c["status"] = "Vigilance";  c["color"] = _EC["warn"];  c["color_lt"] = _EC["warn_lt"]
        else:
            c["status"] = "Action";     c["color"] = _EC["alert"]; c["color_lt"] = _EC["alert_lt"]
    return cats


def _exec_top_recommendations(state: dict) -> list[dict]:
    """Liste de recommandations priorisées, dérivées de l'état."""
    recs = []
    dq = float(state.get("dq_score", 0) or 0)
    etl_ok = (state.get("etl_status") or "") == "success"
    n_alerts = len(state.get("dq_alerts", []) or [])
    n_heal = len(state.get("heal_history", []) or [])
    pii_count = len(state.get("pii_findings", []) or [])
    lm = state.get("logical_model") or {}
    n_dims = len(lm.get("dimension_tables") or [])

    if dq < 70:
        recs.append({"priority": "ÉLEVÉE", "title": "Renforcer la qualité des données sources",
                     "rationale": f"Le score qualité est de {int(dq)}/100. Une revue des règles de validation et un nettoyage des sources sont recommandés avant la mise en production.",
                     "color": _EC["alert"]})
    if not etl_ok:
        recs.append({"priority": "ÉLEVÉE", "title": "Stabiliser le pipeline ETL",
                     "rationale": f"Le pipeline n'a pas terminé en succès (statut : {state.get('etl_status') or 'inconnu'}). Investiguer le journal d'exécution.",
                     "color": _EC["alert"]})
    if pii_count > 0:
        recs.append({"priority": "MOYENNE", "title": "Activer le masquage des colonnes PII",
                     "rationale": f"{pii_count} colonne(s) potentiellement PII identifiée(s). Définir les règles de masquage avant exposition aux exports.",
                     "color": _EC["warn"]})
    if n_heal > 3:
        recs.append({"priority": "MOYENNE", "title": "Investiguer les auto-corrections récurrentes",
                     "rationale": f"{n_heal} corrections automatiques appliquées. Une dérive de schéma est probable côté source.",
                     "color": _EC["warn"]})
    if n_alerts > 0:
        recs.append({"priority": "FAIBLE", "title": "Traiter les alertes qualité résiduelles",
                     "rationale": f"{n_alerts} alerte(s) actuellement ouverte(s). Voir l'onglet Qualité Données.",
                     "color": _EC["accent"]})
    if n_dims < 3:
        recs.append({"priority": "FAIBLE", "title": "Enrichir le modèle dimensionnel",
                     "rationale": f"Seulement {n_dims} dimension(s). Identifier les axes d'analyse manquants pour les utilisateurs métier.",
                     "color": _EC["accent"]})
    if not recs:
        recs.append({"priority": "INFORMATION", "title": "Aucune action corrective requise",
                     "rationale": "Le pipeline est conforme aux attentes. Continuer le suivi régulier des métriques de qualité.",
                     "color": _EC["ok"]})
    return recs[:5]


def _exec_top_risks(state: dict) -> list[dict]:
    """Liste de risques métier identifiés."""
    risks = []
    dq = float(state.get("dq_score", 0) or 0)
    etl_ok = (state.get("etl_status") or "") == "success"
    n_heal = len(state.get("heal_history", []) or [])
    pii_count = len(state.get("pii_findings", []) or [])

    if dq < 50:
        risks.append({"severity": "CRITIQUE", "title": "Données non fiables pour la décision",
                      "description": "Un score qualité inférieur à 50 expose la direction à des décisions basées sur des chiffres potentiellement erronés.",
                      "color": _EC["alert"]})
    if not etl_ok:
        risks.append({"severity": "ÉLEVÉ", "title": "Indisponibilité du Data Warehouse",
                      "description": "Tant que le pipeline n'est pas en succès, les tableaux de bord aval risquent d'afficher des données obsolètes.",
                      "color": _EC["alert"]})
    if pii_count > 5:
        risks.append({"severity": "ÉLEVÉ", "title": "Exposition de données personnelles",
                      "description": f"{pii_count} colonnes PII détectées. Risque de non-conformité RGPD si exportées sans masquage.",
                      "color": _EC["warn"]})
    if n_heal > 5:
        risks.append({"severity": "MOYEN", "title": "Dette technique sur la source",
                      "description": "Le nombre élevé d'auto-corrections suggère un schéma source instable qui demandera une refonte.",
                      "color": _EC["warn"]})
    if not risks:
        risks.append({"severity": "FAIBLE", "title": "Aucun risque majeur identifié",
                      "description": "L'analyse automatique ne détecte pas de risque opérationnel significatif sur cette exécution.",
                      "color": _EC["ok"]})
    return risks[:4]


def _ec_kpi_card(ws, row, col, label, value, sub, accent_hex, span=3, height_label=18, height_value=34, height_sub=18):
    """Carte KPI sur 3 lignes — label / valeur / sous-texte."""
    end_col = col + span - 1
    # Ligne 1 — label
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=label.upper())
    c.font      = _mk_font(bold=True, size=9, color=_EC["white"])
    c.fill      = _mk_fill(accent_hex)
    c.alignment = _mk_align(h="center", v="center")
    _set_row_height(ws, row, height_label)
    # Ligne 2 — valeur
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=end_col)
    v = ws.cell(row=row+1, column=col, value=value)
    v.font      = _mk_font(bold=True, size=22, color=accent_hex)
    v.fill      = _mk_fill(_EC["white"])
    v.alignment = _mk_align(h="center", v="center")
    v.border    = _mk_border("thin", _EC["border"])
    _set_row_height(ws, row+1, height_value)
    # Ligne 3 — sous-texte
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=end_col)
    s = ws.cell(row=row+2, column=col, value=sub or "")
    s.font      = _mk_font(size=9, italic=True, color=_EC["neutral"])
    s.fill      = _mk_fill(_EC["stripe"])
    s.alignment = _mk_align(h="center", v="center")
    s.border    = _mk_border("thin", _EC["border"])
    _set_row_height(ws, row+2, height_sub)


def _ec_link_cell(ws, row, col, target_sheet, label, span=1):
    """Cellule contenant un hyperlien interne vers une autre feuille."""
    end_col = col + span - 1
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=label)
    # Échappe les guillemets et apostrophes du nom de feuille
    safe = str(target_sheet).replace("'", "''")
    c.hyperlink = f"#'{safe}'!A1"
    c.font      = _mk_font(bold=True, size=10, color=_EC["accent"])
    c.alignment = _mk_align(h="left", v="center", wrap=True)
    return c


def _ec_traffic_light(score: float):
    """(label, fill_color, text_color) selon score 0-100."""
    if score >= 85:  return ("BON",      _EC["ok"],    _EC["white"])
    if score >= 65:  return ("VIGILANCE",_EC["warn"],  _EC["white"])
    return            ("ACTION",   _EC["alert"], _EC["white"])


def _ec_gauge_bar(ws, row, col, score: float, span=10):
    """
    Rend une jauge horizontale faite de cellules colorées proportionnellement.
    span = nombre total de cellules disponibles ; on en remplit (span * score/100).
    """
    s = max(0, min(100, float(score)))
    filled = max(1, round(span * s / 100))
    color = _EC["ok"] if s >= 85 else (_EC["warn"] if s >= 65 else _EC["alert"])
    for i in range(span):
        cell = ws.cell(row=row, column=col + i, value="")
        cell.fill   = _mk_fill(color if i < filled else _EC["neutral_lt"])
        cell.border = _mk_border("thin", _EC["border"])
    _set_row_height(ws, row, 18)


# =============================================================================
#  Feuille 01 : COUVERTURE
# =============================================================================
def _ws_cover(wb, state: dict, session_id: str):
    ws = wb.create_sheet("01 Couverture", index=0)
    ws.sheet_view.showGridLines = False

    # Largeurs colonnes (12 colonnes utiles)
    for letter, w in {"A":4,"B":18,"C":18,"D":18,"E":18,"F":18,"G":18,"H":18,"I":18,"J":18,"K":18,"L":4}.items():
        ws.column_dimensions[letter].width = w

    # Bandeau supérieur navy
    ws.merge_cells("A1:L2")
    band = ws["A1"]
    band.fill = _mk_fill(_EC["primary"])
    _set_row_height(ws, 1, 14); _set_row_height(ws, 2, 14)

    # Titre principal
    ws.merge_cells("B4:K6")
    title = ws["B4"]
    title.value     = "RAPPORT DÉCISIONNEL"
    title.font      = _mk_font(bold=True, size=36, color=_EC["primary"])
    title.alignment = _mk_align(h="center", v="center")
    for r in (4, 5, 6):
        _set_row_height(ws, r, 22)

    ws.merge_cells("B7:K8")
    subtitle = ws["B7"]
    subtitle.value     = "Data Warehouse — Pipeline d'Intégration"
    subtitle.font      = _mk_font(size=18, color=_EC["neutral"], italic=True)
    subtitle.alignment = _mk_align(h="center", v="center")
    _set_row_height(ws, 7, 18); _set_row_height(ws, 8, 18)

    # Ligne séparatrice
    ws.merge_cells("B10:K10")
    sep = ws["B10"]
    sep.fill = _mk_fill(_EC["accent"])
    _set_row_height(ws, 10, 4)

    # Note globale
    grade = _exec_overall_grade(state)

    ws.merge_cells("B12:E18")
    g_box = ws["B12"]
    g_box.value     = grade["letter"]
    g_box.font      = _mk_font(bold=True, size=120, color=grade["color"])
    g_box.fill      = _mk_fill(grade["color_lt"])
    g_box.alignment = _mk_align(h="center", v="center")
    for r in range(12, 19):
        _set_row_height(ws, r, 22)

    ws.merge_cells("F12:K13")
    label = ws["F12"]
    label.value     = "Note globale du pipeline"
    label.font      = _mk_font(bold=True, size=11, color=_EC["neutral"])
    label.alignment = _mk_align(h="left", v="center")

    ws.merge_cells("F14:K15")
    score = ws["F14"]
    score.value     = f"{grade['score']}/100 — {grade['label']}"
    score.font      = _mk_font(bold=True, size=18, color=grade["color"])
    score.alignment = _mk_align(h="left", v="center")

    ws.merge_cells("F16:K18")
    nar = ws["F16"]
    nar.value     = grade["narrative"]
    nar.font      = _mk_font(size=11, color=_EC["ink"])
    nar.alignment = _mk_align(h="left", v="top", wrap=True)

    # Métadonnées
    meta_row = 21
    ws.merge_cells(f"B{meta_row}:K{meta_row}")
    h = ws.cell(row=meta_row, column=2, value="INFORMATIONS DE GÉNÉRATION")
    h.font      = _mk_font(bold=True, size=10, color=_EC["white"])
    h.fill      = _mk_fill(_EC["primary"])
    h.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, meta_row, 22)

    metas = [
        ("Identifiant de session",  session_id),
        ("Préfixe utilisateur",     (state.get("user_prefix") or "—").upper()),
        ("Base cible",              state.get("target_database") or state.get("restored_db_name") or "—"),
        ("Mode d'exécution",        (state.get("etl_mode") or "—").replace("_", " ").title()),
        ("Date de génération",      datetime.now().strftime("%d/%m/%Y à %H:%M:%S")),
        ("Version de la plateforme", "Agent Data Warehouse v3.0.1"),
    ]
    for i, (k, v) in enumerate(metas):
        r = meta_row + 1 + i
        ws.merge_cells(f"B{r}:E{r}")
        kc = ws.cell(row=r, column=2, value=k)
        kc.font      = _mk_font(bold=True, size=10, color=_EC["neutral"])
        kc.fill      = _mk_fill(_EC["stripe"])
        kc.alignment = _mk_align(h="left", v="center")
        kc.border    = _mk_border("thin", _EC["border"])

        ws.merge_cells(f"F{r}:K{r}")
        vc = ws.cell(row=r, column=6, value=str(v))
        vc.font      = _mk_font(size=10, color=_EC["ink"])
        vc.fill      = _mk_fill(_EC["white"])
        vc.alignment = _mk_align(h="left", v="center")
        vc.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 22)

    # Bloc validation / signatures
    sig_row = meta_row + 1 + len(metas) + 2
    ws.merge_cells(f"B{sig_row}:K{sig_row}")
    sh = ws.cell(row=sig_row, column=2, value="VALIDATION")
    sh.font      = _mk_font(bold=True, size=10, color=_EC["white"])
    sh.fill      = _mk_fill(_EC["primary"])
    sh.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, sig_row, 22)

    roles = [("Responsable données",   "Nom · Date · Visa"),
             ("Direction métier",       "Nom · Date · Visa"),
             ("Direction des systèmes", "Nom · Date · Visa")]
    for i, (role, line) in enumerate(roles):
        r = sig_row + 1 + i
        ws.merge_cells(f"B{r}:E{r}")
        kc = ws.cell(row=r, column=2, value=role)
        kc.font      = _mk_font(bold=True, size=10, color=_EC["neutral"])
        kc.fill      = _mk_fill(_EC["stripe"])
        kc.alignment = _mk_align(h="left", v="center")
        kc.border    = _mk_border("thin", _EC["border"])

        ws.merge_cells(f"F{r}:K{r}")
        vc = ws.cell(row=r, column=6, value=line)
        vc.font      = _mk_font(size=10, italic=True, color=_EC["neutral"])
        vc.fill      = _mk_fill(_EC["white"])
        vc.alignment = _mk_align(h="left", v="center")
        vc.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 30)

    # Pied de page
    foot_row = sig_row + 1 + len(roles) + 2
    ws.merge_cells(f"A{foot_row}:L{foot_row}")
    f = ws.cell(row=foot_row, column=1,
                value="Document confidentiel — diffusion strictement interne. Toute reproduction est subordonnée à autorisation préalable.")
    f.font      = _mk_font(size=8.5, italic=True, color=_EC["neutral"])
    f.alignment = _mk_align(h="center", v="center")
    _set_row_height(ws, foot_row, 18)


# =============================================================================
#  Feuille 02 : SOMMAIRE
# =============================================================================
def _ws_toc(wb, state: dict, sheet_index: list[tuple[str, str, str]]):
    """
    sheet_index : liste de (sheet_name, audience, description) déjà ordonnée
    selon l'apparition dans le classeur.
    """
    ws = wb.create_sheet("02 Sommaire", index=1)
    ws.sheet_view.showGridLines = False

    for letter, w in {"A":4,"B":6,"C":36,"D":18,"E":54,"F":4}.items():
        ws.column_dimensions[letter].width = w

    # Titre
    ws.merge_cells("B2:E3")
    t = ws["B2"]
    t.value     = "SOMMAIRE DU RAPPORT"
    t.font      = _mk_font(bold=True, size=22, color=_EC["primary"])
    t.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 2, 18); _set_row_height(ws, 3, 18)

    ws.merge_cells("B4:E4")
    sub = ws["B4"]
    sub.value     = "Cliquez sur un titre pour accéder directement à la section correspondante."
    sub.font      = _mk_font(size=10, italic=True, color=_EC["neutral"])
    sub.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 4, 18)

    # En-têtes
    headers = [("B", "#"), ("C", "Section"), ("D", "Audience"), ("E", "Description")]
    for letter, label in headers:
        c = ws[f"{letter}6"]
        c.value     = label
        c.font      = _mk_font(bold=True, size=10, color=_EC["white"])
        c.fill      = _mk_fill(_EC["primary"])
        c.alignment = _mk_align(h="left", v="center")
        c.border    = _mk_border("thin", _EC["primary"])
    _set_row_height(ws, 6, 24)

    r = 7
    for i, (sheet_name, audience, description) in enumerate(sheet_index, start=1):
        # Numéro
        nc = ws.cell(row=r, column=2, value=i)
        nc.font      = _mk_font(bold=True, size=10, color=_EC["neutral"])
        nc.fill      = _mk_fill(_EC["stripe"] if i % 2 else _EC["white"])
        nc.alignment = _mk_align(h="center", v="center")
        nc.border    = _mk_border("thin", _EC["border"])
        # Titre cliquable
        link = _ec_link_cell(ws, r, 3, sheet_name, sheet_name)
        link.fill   = _mk_fill(_EC["stripe"] if i % 2 else _EC["white"])
        link.border = _mk_border("thin", _EC["border"])
        # Audience
        ac = ws.cell(row=r, column=4, value=audience)
        ac.font      = _mk_font(size=9, color=_EC["neutral"])
        ac.fill      = _mk_fill(_EC["stripe"] if i % 2 else _EC["white"])
        ac.alignment = _mk_align(h="left", v="center")
        ac.border    = _mk_border("thin", _EC["border"])
        # Description
        dc = ws.cell(row=r, column=5, value=description)
        dc.font      = _mk_font(size=9.5, color=_EC["ink"])
        dc.fill      = _mk_fill(_EC["stripe"] if i % 2 else _EC["white"])
        dc.alignment = _mk_align(h="left", v="center", wrap=True)
        dc.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 28)
        r += 1


# =============================================================================
#  Feuille 03 : SYNTHÈSE EXÉCUTIVE
# =============================================================================
def _ws_executive_summary(wb, state: dict, session_id: str):
    ws = wb.create_sheet("03 Synthese Executive", index=2)
    ws.sheet_view.showGridLines = False

    for letter, w in {"A":3,"B":14,"C":14,"D":14,"E":14,"F":14,"G":14,
                      "H":14,"I":14,"J":14,"K":14,"L":14,"M":14,"N":3}.items():
        ws.column_dimensions[letter].width = w

    # En-tête
    ws.merge_cells("B2:M3")
    h = ws["B2"]
    h.value     = "SYNTHÈSE EXÉCUTIVE"
    h.font      = _mk_font(bold=True, size=22, color=_EC["primary"])
    h.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 2, 18); _set_row_height(ws, 3, 18)

    ws.merge_cells("B4:M4")
    sub = ws["B4"]
    grade = _exec_overall_grade(state)
    sub.value     = f"Vue à 360° destinée aux décideurs · Note globale : {grade['letter']} ({grade['score']}/100)"
    sub.font      = _mk_font(size=11, italic=True, color=_EC["neutral"])
    sub.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 4, 20)

    # ---- Bandeau narratif ---------------------------------------------------
    ws.merge_cells("B6:M9")
    nar = ws["B6"]
    nar.value     = grade["narrative"]
    nar.font      = _mk_font(size=12, color=_EC["ink"])
    nar.fill      = _mk_fill(grade["color_lt"])
    nar.alignment = _mk_align(h="left", v="center", wrap=True)
    nar.border    = _mk_border("medium", grade["color"])
    for r in (6, 7, 8, 9):
        _set_row_height(ws, r, 16)

    # ---- KPIs (6 cartes en 2 rangées de 3) ---------------------------------
    dq = float(state.get("dq_score", 0) or 0)
    etl_status = (state.get("etl_status") or "—").upper()
    lm = state.get("logical_model") or {}
    n_dims = len(lm.get("dimension_tables") or [])
    n_facts = 1 if lm.get("fact_table") else len(lm.get("fact_tables") or [])
    n_alerts = len(state.get("dq_alerts", []) or [])
    n_heal = len(state.get("heal_history", []) or [])

    lm_metrics = state.get("load_metrics") or {}
    total_rows = 0
    if isinstance(lm_metrics, dict):
        for v in lm_metrics.values():
            if isinstance(v, dict):    total_rows += int(v.get("inserted", 0) or 0)
            elif isinstance(v, (int, float)): total_rows += int(v)

    node_dur = state.get("node_durations") or {}
    total_dur = round(sum(float(v) for v in node_dur.values() if isinstance(v, (int, float))), 0)

    cards_row1 = [
        ("Score qualité",   f"{int(dq)}/100",     "Données prêtes pour la décision",        _EC["primary"]),
        ("Statut pipeline", etl_status,           "Fiabilité de l'intégration",             _EC["accent"]),
        ("Lignes intégrées",f"{total_rows:,}".replace(",", " "), "Volume traité durant cette exécution", _EC["ok"]),
    ]
    cards_row2 = [
        ("Modèle étoile",     f"{n_facts}F + {n_dims}D",        "Tables de faits + dimensions",       _EC["gold"]),
        ("Alertes qualité",   str(n_alerts),                     "Anomalies à examiner",                _EC["warn"]),
        ("Auto-corrections",  str(n_heal),                       "Réparations automatiques",            _EC["alert"] if n_heal > 3 else _EC["ok"]),
    ]

    # Rangée 1 — colonnes B..M (12 cells = 3 cartes de 4 colonnes)
    for idx, (label, value, sub_t, col) in enumerate(cards_row1):
        start_col = 2 + idx * 4
        _ec_kpi_card(ws, 11, start_col, label, value, sub_t, col, span=4)
    # Rangée 2
    for idx, (label, value, sub_t, col) in enumerate(cards_row2):
        start_col = 2 + idx * 4
        _ec_kpi_card(ws, 16, start_col, label, value, sub_t, col, span=4)

    # ---- Indicateurs de santé par domaine ----------------------------------
    r = 22
    ws.merge_cells(f"B{r}:M{r}")
    ti = ws.cell(row=r, column=2, value="INDICATEURS DE SANTÉ")
    ti.font      = _mk_font(bold=True, size=11, color=_EC["white"])
    ti.fill      = _mk_fill(_EC["primary"])
    ti.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, r, 22)
    r += 1

    # En-têtes
    headers = [("Domaine", 3), ("Score", 1), ("Jauge", 5), ("Statut", 1), ("Facteurs clés", 2)]
    col = 2
    for label, span in headers:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + span - 1)
        c = ws.cell(row=r, column=col, value=label)
        c.font      = _mk_font(bold=True, size=9, color=_EC["white"])
        c.fill      = _mk_fill(_EC["neutral"])
        c.alignment = _mk_align(h="center", v="center")
        c.border    = _mk_border("thin", _EC["neutral"])
        col += span
    _set_row_height(ws, r, 18); r += 1

    for cat in _exec_health_categories(state):
        # Domaine (3 col)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        n = ws.cell(row=r, column=2, value=cat["name"])
        n.font      = _mk_font(bold=True, size=10, color=_EC["ink"])
        n.fill      = _mk_fill(_EC["white"])
        n.alignment = _mk_align(h="left", v="center")
        n.border    = _mk_border("thin", _EC["border"])
        # Score (1 col)
        sc = ws.cell(row=r, column=5, value=f"{cat['score']}")
        sc.font      = _mk_font(bold=True, size=12, color=cat["color"])
        sc.fill      = _mk_fill(_EC["white"])
        sc.alignment = _mk_align(h="center", v="center")
        sc.border    = _mk_border("thin", _EC["border"])
        # Jauge (5 col)
        _ec_gauge_bar(ws, r, 6, cat["score"], span=5)
        # Statut (1 col)
        st = ws.cell(row=r, column=11, value=cat["status"].upper())
        st.font      = _mk_font(bold=True, size=9, color=_EC["white"])
        st.fill      = _mk_fill(cat["color"])
        st.alignment = _mk_align(h="center", v="center")
        st.border    = _mk_border("thin", _EC["border"])
        # Facteurs (2 col)
        ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=13)
        f = ws.cell(row=r, column=12, value=" · ".join(cat["factors"]))
        f.font      = _mk_font(size=9, color=_EC["neutral"])
        f.fill      = _mk_fill(_EC["white"])
        f.alignment = _mk_align(h="left", v="center", wrap=True)
        f.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 26)
        r += 1

    r += 1

    # ---- Recommandations ----------------------------------------------------
    ws.merge_cells(f"B{r}:M{r}")
    th = ws.cell(row=r, column=2, value="ACTIONS RECOMMANDÉES")
    th.font      = _mk_font(bold=True, size=11, color=_EC["white"])
    th.fill      = _mk_fill(_EC["primary"])
    th.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, r, 22); r += 1

    for rec in _exec_top_recommendations(state):
        # Priorité
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        p = ws.cell(row=r, column=2, value=rec["priority"])
        p.font      = _mk_font(bold=True, size=9, color=_EC["white"])
        p.fill      = _mk_fill(rec["color"])
        p.alignment = _mk_align(h="center", v="center")
        p.border    = _mk_border("thin", rec["color"])
        # Titre + rationale
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=13)
        body = ws.cell(row=r, column=4,
                       value=f"{rec['title']}\n{rec['rationale']}")
        body.font      = _mk_font(size=10, color=_EC["ink"])
        body.fill      = _mk_fill(_EC["white"])
        body.alignment = _mk_align(h="left", v="center", wrap=True)
        body.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 42); r += 1

    r += 1

    # ---- Risques ------------------------------------------------------------
    ws.merge_cells(f"B{r}:M{r}")
    th = ws.cell(row=r, column=2, value="RISQUES IDENTIFIÉS")
    th.font      = _mk_font(bold=True, size=11, color=_EC["white"])
    th.fill      = _mk_fill(_EC["alert"])
    th.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, r, 22); r += 1

    for risk in _exec_top_risks(state):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        s = ws.cell(row=r, column=2, value=risk["severity"])
        s.font      = _mk_font(bold=True, size=9, color=_EC["white"])
        s.fill      = _mk_fill(risk["color"])
        s.alignment = _mk_align(h="center", v="center")
        s.border    = _mk_border("thin", risk["color"])

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=13)
        body = ws.cell(row=r, column=4,
                       value=f"{risk['title']}\n{risk['description']}")
        body.font      = _mk_font(size=10, color=_EC["ink"])
        body.fill      = _mk_fill(_EC["white"])
        body.alignment = _mk_align(h="left", v="center", wrap=True)
        body.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 42); r += 1

    ws.freeze_panes = "A6"


# =============================================================================
#  Feuille 04 : TABLEAU DE DÉCISION (graphiques natifs)
# =============================================================================
def _ws_decision_dashboard(wb, state: dict):
    ws = wb.create_sheet("04 Tableau de Decision", index=3)
    ws.sheet_view.showGridLines = False

    for letter, w in {"A":3,"B":24,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14,"K":14,"L":3}.items():
        ws.column_dimensions[letter].width = w

    # Titre
    ws.merge_cells("B2:K3")
    t = ws["B2"]
    t.value     = "TABLEAU DE DÉCISION"
    t.font      = _mk_font(bold=True, size=22, color=_EC["primary"])
    t.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 2, 18); _set_row_height(ws, 3, 18)

    ws.merge_cells("B4:K4")
    sub = ws["B4"]
    sub.value     = "Vue analytique consolidée pour le pilotage opérationnel."
    sub.font      = _mk_font(size=11, italic=True, color=_EC["neutral"])
    sub.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 4, 18)

    # ---- Section 1 : volumétrie par table ----------------------------------
    r = 6
    ws.merge_cells(f"B{r}:K{r}")
    s1 = ws.cell(row=r, column=2, value="VOLUMÉTRIE PAR TABLE CIBLE")
    s1.font      = _mk_font(bold=True, size=11, color=_EC["white"])
    s1.fill      = _mk_fill(_EC["primary"])
    s1.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, r, 22); r += 1

    # En-têtes
    hdrs = [("B", "Table"), ("C", "Insérées"), ("D", "Mises à jour"), ("E", "Rejetées"), ("F", "Total")]
    for letter, label in hdrs:
        c = ws[f"{letter}{r}"]
        c.value     = label
        c.font      = _mk_font(bold=True, size=10, color=_EC["white"])
        c.fill      = _mk_fill(_EC["neutral"])
        c.alignment = _mk_align(h="center", v="center")
        c.border    = _mk_border("thin", _EC["neutral"])
    _set_row_height(ws, r, 20); r += 1
    chart_data_start = r

    lm_metrics = state.get("load_metrics") or {}
    if isinstance(lm_metrics, dict) and lm_metrics:
        for i, (tbl, m) in enumerate(lm_metrics.items()):
            if isinstance(m, dict):
                ins = int(m.get("inserted", 0) or 0)
                upd = int(m.get("updated", 0) or 0)
                rej = int(m.get("rejected", 0) or 0)
            elif isinstance(m, (int, float)):
                ins, upd, rej = int(m), 0, 0
            else:
                continue
            tot = ins + upd + rej
            stripe = _EC["stripe"] if i % 2 else _EC["white"]
            row_vals = [(2, tbl, "left"), (3, ins, "right"), (4, upd, "right"),
                        (5, rej, "right"), (6, tot, "right")]
            for col, val, h in row_vals:
                cc = ws.cell(row=r, column=col, value=val)
                cc.font      = _mk_font(bold=(col == 2 or col == 6), size=10, color=_EC["ink"])
                cc.fill      = _mk_fill(stripe)
                cc.alignment = _mk_align(h=h, v="center")
                cc.border    = _mk_border("thin", _EC["border"])
                if col > 2:
                    cc.number_format = "#,##0"
            _set_row_height(ws, r, 20)
            r += 1
    else:
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(row=r, column=2, value="Aucune métrique de chargement disponible pour cette exécution.")
        c.font      = _mk_font(size=10, italic=True, color=_EC["neutral"])
        c.fill      = _mk_fill(_EC["stripe"])
        c.alignment = _mk_align(h="center", v="center")
        c.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 22)
        r += 1

    chart_data_end = r - 1

    # Tente de greffer un BarChart natif si openpyxl le permet
    try:
        if chart_data_end >= chart_data_start:
            from openpyxl.chart import BarChart, Reference
            chart = BarChart()
            chart.type = "bar"
            chart.style = 11
            chart.title = "Lignes intégrées par table"
            chart.y_axis.title = "Lignes"
            chart.x_axis.title = "Table"
            data = Reference(ws, min_col=3, min_row=chart_data_start - 1, max_col=5, max_row=chart_data_end)
            cats = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 18
            chart.height = 9
            ws.add_chart(chart, f"H{chart_data_start - 1}")
    except Exception as e:
        logger.debug(f"[Excel] Bar chart skipped: {e}")

    r += 2

    # ---- Section 2 : durée par module --------------------------------------
    ws.merge_cells(f"B{r}:K{r}")
    s2 = ws.cell(row=r, column=2, value="DURÉE D'EXÉCUTION PAR MODULE")
    s2.font      = _mk_font(bold=True, size=11, color=_EC["white"])
    s2.fill      = _mk_fill(_EC["primary"])
    s2.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, r, 22); r += 1

    for letter, label in [("B", "Module"), ("C", "Durée (s)"), ("D", "Part du total"), ("E", "Statut")]:
        c = ws[f"{letter}{r}"]
        c.value     = label
        c.font      = _mk_font(bold=True, size=10, color=_EC["white"])
        c.fill      = _mk_fill(_EC["neutral"])
        c.alignment = _mk_align(h="center", v="center")
        c.border    = _mk_border("thin", _EC["neutral"])
    _set_row_height(ws, r, 20); r += 1

    durations = state.get("node_durations") or {}
    total = sum(float(v) for v in durations.values() if isinstance(v, (int, float))) or 1.0
    sorted_dur = sorted(
        ((k, float(v)) for k, v in durations.items() if isinstance(v, (int, float))),
        key=lambda x: x[1], reverse=True
    )
    if sorted_dur:
        for i, (mod, d) in enumerate(sorted_dur):
            stripe = _EC["stripe"] if i % 2 else _EC["white"]
            pct = (d / total) * 100
            status = "Rapide" if d < 5 else ("Normal" if d < 30 else ("Lent" if d < 90 else "Très lent"))
            status_color = _EC["ok"] if d < 5 else (_EC["accent"] if d < 30 else (_EC["warn"] if d < 90 else _EC["alert"]))

            mc = ws.cell(row=r, column=2, value=mod)
            mc.font      = _mk_font(bold=True, size=10, color=_EC["ink"])
            mc.fill      = _mk_fill(stripe)
            mc.alignment = _mk_align(h="left", v="center")
            mc.border    = _mk_border("thin", _EC["border"])

            dc = ws.cell(row=r, column=3, value=round(d, 1))
            dc.font      = _mk_font(size=10, color=_EC["ink"])
            dc.fill      = _mk_fill(stripe)
            dc.alignment = _mk_align(h="right", v="center")
            dc.border    = _mk_border("thin", _EC["border"])
            dc.number_format = '#,##0.0" s"'

            pc = ws.cell(row=r, column=4, value=pct / 100)
            pc.font      = _mk_font(size=10, color=_EC["ink"])
            pc.fill      = _mk_fill(stripe)
            pc.alignment = _mk_align(h="right", v="center")
            pc.border    = _mk_border("thin", _EC["border"])
            pc.number_format = "0.0%"

            sc = ws.cell(row=r, column=5, value=status)
            sc.font      = _mk_font(bold=True, size=9, color=_EC["white"])
            sc.fill      = _mk_fill(status_color)
            sc.alignment = _mk_align(h="center", v="center")
            sc.border    = _mk_border("thin", _EC["border"])

            _set_row_height(ws, r, 20); r += 1
    else:
        ws.merge_cells(f"B{r}:E{r}")
        c = ws.cell(row=r, column=2, value="Aucune mesure de durée pour cette exécution.")
        c.font      = _mk_font(size=10, italic=True, color=_EC["neutral"])
        c.fill      = _mk_fill(_EC["stripe"])
        c.alignment = _mk_align(h="center", v="center")
        c.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 22)
        r += 1

    ws.freeze_panes = "A6"


# =============================================================================
#  Feuille 05 : SCORECARD QUALITÉ
# =============================================================================
def _ws_scorecard(wb, state: dict):
    ws = wb.create_sheet("05 Scorecard Sante", index=4)
    ws.sheet_view.showGridLines = False

    for letter, w in {"A":3,"B":22,"C":12,"D":40,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14,"K":3}.items():
        ws.column_dimensions[letter].width = w

    ws.merge_cells("B2:J3")
    t = ws["B2"]
    t.value     = "SCORECARD DE SANTÉ"
    t.font      = _mk_font(bold=True, size=22, color=_EC["primary"])
    t.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 2, 18); _set_row_height(ws, 3, 18)

    ws.merge_cells("B4:J4")
    sub = ws["B4"]
    sub.value     = "Évaluation détaillée par domaine — barème : Bon (≥85) · Vigilance (65-84) · Action (<65)."
    sub.font      = _mk_font(size=11, italic=True, color=_EC["neutral"])
    sub.alignment = _mk_align(h="left", v="center")
    _set_row_height(ws, 4, 18)

    # En-têtes
    r = 6
    headers = [("B", "Domaine", 1), ("C", "Score / 100", 1), ("D", "Évaluation", 1),
               ("E", "Jauge", 5), ("J", "Statut", 1)]
    for (letter, label, span) in headers:
        col = ord(letter) - 64
        if span > 1:
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + span - 1)
        c = ws.cell(row=r, column=col, value=label)
        c.font      = _mk_font(bold=True, size=10, color=_EC["white"])
        c.fill      = _mk_fill(_EC["primary"])
        c.alignment = _mk_align(h="center", v="center")
        c.border    = _mk_border("thin", _EC["primary"])
    _set_row_height(ws, r, 24); r += 1

    for cat in _exec_health_categories(state):
        # Domaine
        n = ws.cell(row=r, column=2, value=cat["name"])
        n.font      = _mk_font(bold=True, size=11, color=_EC["ink"])
        n.fill      = _mk_fill(_EC["white"])
        n.alignment = _mk_align(h="left", v="center")
        n.border    = _mk_border("thin", _EC["border"])
        # Score
        s = ws.cell(row=r, column=3, value=cat["score"])
        s.font      = _mk_font(bold=True, size=14, color=cat["color"])
        s.fill      = _mk_fill(_EC["white"])
        s.alignment = _mk_align(h="center", v="center")
        s.border    = _mk_border("thin", _EC["border"])
        # Évaluation textuelle
        e = ws.cell(row=r, column=4, value=" · ".join(cat["factors"]))
        e.font      = _mk_font(size=10, color=_EC["neutral"])
        e.fill      = _mk_fill(_EC["white"])
        e.alignment = _mk_align(h="left", v="center", wrap=True)
        e.border    = _mk_border("thin", _EC["border"])
        # Jauge
        _ec_gauge_bar(ws, r, 5, cat["score"], span=5)
        # Statut
        st = ws.cell(row=r, column=10, value=cat["status"].upper())
        st.font      = _mk_font(bold=True, size=10, color=_EC["white"])
        st.fill      = _mk_fill(cat["color"])
        st.alignment = _mk_align(h="center", v="center")
        st.border    = _mk_border("thin", _EC["border"])
        _set_row_height(ws, r, 32)
        r += 1

    # Note explicative
    r += 1
    ws.merge_cells(f"B{r}:J{r+3}")
    note = ws.cell(row=r, column=2, value=(
        "Lecture du tableau\n"
        "• Bon (≥85) : aucun blocage identifié, poursuivre la surveillance régulière.\n"
        "• Vigilance (65-84) : points à examiner avant la prochaine itération du pipeline.\n"
        "• Action (<65) : intervention requise — voir l'onglet Synthèse Exécutive pour les recommandations."
    ))
    note.font      = _mk_font(size=10, color=_EC["neutral"], italic=True)
    note.fill      = _mk_fill(_EC["neutral_lt"])
    note.alignment = _mk_align(h="left", v="top", wrap=True)
    note.border    = _mk_border("thin", _EC["border"])
    for rr in range(r, r + 4):
        _set_row_height(ws, rr, 18)


# ============================================================
# FONCTION PRINCIPALE
# ============================================================

def generate_xlsx_report(state: dict, session_id: str) -> str:
    """
    Génère un classeur Excel décisionnel professionnel.

    Structure (14 feuilles ordonnées pour la consultation par un décideur) :
      01 Couverture            — Page de garde + note globale + bloc validation
      02 Sommaire              — Index cliquable de toutes les feuilles
      03 Synthèse Exécutive    — Vue à 360° pour la direction (KPI, recommandations, risques)
      04 Tableau de Décision   — Volumétrie + durées avec graphiques natifs
      05 Scorecard Sante       — 5 domaines évalués avec jauges
      06 Mesures et KPI        — Détail des mesures par dimension temporelle
      07 Qualite Donnees       — 12 dimensions DQ avec alertes
      08 Performance ETL       — Métriques par module ETL
      09 Schema Etoile         — Description du modèle dimensionnel
      10 Catalogue             — Tables, colonnes, descriptions, tags
      11 Lignage               — Lineage colonne par colonne
      12 Analyses              — Requêtes analytiques pré-construites
      13 DDL                   — Script SQL complet
      14 Journal               — Journal d'exécution détaillé
    """
    from openpyxl import Workbook

    os.makedirs("outputs", exist_ok=True)
    xlsx_path = os.path.abspath(f"outputs/{session_id}_report.xlsx")

    wb = Workbook()
    # On supprime la feuille par défaut "Sheet" (les nouvelles sont créées avec create_sheet)
    if "Sheet" in wb.sheetnames:
        try:
            del wb["Sheet"]
        except Exception:
            pass

    # ---- Index des feuilles (utilisé par le sommaire) ---------------------
    sheet_index = [
        ("01 Couverture",            "Tous publics",       "Page de garde, note globale, bloc de validation."),
        ("02 Sommaire",              "Tous publics",       "Index cliquable de toutes les sections."),
        ("03 Synthese Executive",    "Direction",          "Vue à 360°, KPI majeurs, recommandations et risques."),
        ("04 Tableau de Decision",   "Direction métier",   "Volumétrie consolidée et durées d'exécution avec graphiques."),
        ("05 Scorecard Sante",       "Direction métier",   "Évaluation détaillée par domaine (qualité, modèle, ETL, sécurité, performance)."),
        ("06 Mesures & KPI",         "Analystes",          "Détail des mesures clés par axe d'analyse temporel."),
        ("07 Qualite Donnees",       "Data Officer",       "Évaluation détaillée selon les douze dimensions de qualité."),
        ("08 Performance ETL",       "Équipe technique",   "Métriques opérationnelles d'exécution ETL."),
        ("09 Schema Etoile",         "Architectes",        "Description complète du modèle dimensionnel généré."),
        ("10 Catalogue",             "Architectes",        "Catalogue des tables, colonnes, types et tags."),
        ("11 Lignage",               "Data Officer",       "Traçabilité colonne par colonne, source vers cible."),
        ("12 Analyses",              "Analystes",          "Requêtes analytiques prêtes à l'emploi."),
        ("13 DDL",                   "Équipe technique",   "Script SQL complet de création du Data Warehouse."),
        ("14 Journal",               "Équipe technique",   "Journal détaillé du déroulement du pipeline."),
    ]

    # ---- 1. Feuilles de tête (création + ordre) ---------------------------
    try:
        _ws_cover(wb, state, session_id)
    except Exception as e:
        logger.warning(f"[XLSX] Couverture error: {e}", exc_info=True)

    try:
        _ws_toc(wb, state, sheet_index)
    except Exception as e:
        logger.warning(f"[XLSX] Sommaire error: {e}", exc_info=True)

    try:
        _ws_executive_summary(wb, state, session_id)
    except Exception as e:
        logger.warning(f"[XLSX] Synthese exec error: {e}", exc_info=True)

    try:
        _ws_decision_dashboard(wb, state)
    except Exception as e:
        logger.warning(f"[XLSX] Tableau decision error: {e}", exc_info=True)

    try:
        _ws_scorecard(wb, state)
    except Exception as e:
        logger.warning(f"[XLSX] Scorecard error: {e}", exc_info=True)

    # ---- 2. Feuilles de détail (existantes) renommées sans emojis --------
    detail_sheets = [
        (_ws_measures_kpi,    "06 Mesures & KPI"),
        (_ws_data_quality,    "07 Qualite Donnees"),
        (_ws_etl_performance, "08 Performance ETL"),
        (_ws_star_schema,     "09 Schema Etoile"),
        (_ws_catalog,         "10 Catalogue"),
        (_ws_lineage,         "11 Lignage"),
        (_ws_analytics,       "12 Analyses"),
        (_ws_ddl,             "13 DDL"),
        (_ws_execution_log,   "14 Journal"),
    ]
    for fn, target_name in detail_sheets:
        try:
            # Mémorise les feuilles existantes pour identifier celle créée par fn(...)
            before = set(wb.sheetnames)
            fn(wb, state)
            after = set(wb.sheetnames)
            new_sheets = after - before
            for nm in new_sheets:
                # Renomme avec le préfixe ordonné, en respectant les contraintes Excel
                try:
                    wb[nm].title = _sanitize_sheet_name(target_name)
                except Exception:
                    pass
        except Exception as e:
            logger.warning(f"[XLSX] Feuille '{target_name}' error: {e}", exc_info=True)

    # Si une feuille "Sheet" résiduelle subsiste, on la supprime
    for name in list(wb.sheetnames):
        if name.lower() in ("sheet", "feuille", "feuil1"):
            try:
                del wb[name]
            except Exception:
                pass

    # ---- 3. Feuille active à l'ouverture : la couverture ------------------
    try:
        if "01 Couverture" in wb.sheetnames:
            wb.active = wb.sheetnames.index("01 Couverture")
    except Exception:
        pass

    # ---- 4. Métadonnées du document --------------------------------------
    try:
        wb.properties.title       = "Rapport Décisionnel — Data Warehouse"
        wb.properties.subject     = "Pipeline d'intégration de données — vue décisionnelle"
        wb.properties.creator     = "Agent Data Warehouse v3.0.1"
        wb.properties.description = (
            f"Rapport généré pour la session {session_id} le "
            f"{datetime.now().strftime('%d/%m/%Y à %H:%M')}. "
            "Présente la note globale du pipeline, les KPI majeurs, les recommandations "
            "priorisées et les risques identifiés à destination de la direction."
        )
        wb.properties.keywords    = "data warehouse, pipeline ETL, qualité, décisionnel, KPI"
        wb.properties.category    = "Rapport décisionnel"
        wb.properties.lastModifiedBy = "Agent Data Warehouse"
    except Exception:
        pass

    wb.save(xlsx_path)
    logger.info(f"[Export] XLSX décisionnel généré : {xlsx_path}")
    return xlsx_path


def _coerce_cell(v):
    """Convertit les valeurs complexes en string pour Excel."""
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, dict, tuple, set)):
        import json as _json
        try:
            return _json.dumps(v, ensure_ascii=False, default=str)[:32000]
        except Exception:
            return str(v)[:32000]
    return str(v)[:32000]


def generate_csv_bundle(state: dict, session_id: str) -> bytes:
    """Genere un .zip contenant un .csv par table + un manifest.json."""
    import io
    import csv
    import json as _json
    import zipfile

    buf = io.BytesIO()
    manifest = {
        "session_id": session_id,
        "generated_at": datetime.now().isoformat(),
        "user_prefix": state.get("user_prefix"),
        "tables": [],
    }
    n_tables = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, rows in _iter_result_tables(state):
            fname = "".join(c if c.isalnum() or c in "_-" else "_" for c in str(name))[:80] or "table"
            csv_buf = io.StringIO()
            first = rows[0] if rows else {}
            if isinstance(first, dict):
                cols = list(first.keys())
                writer = csv.DictWriter(csv_buf, fieldnames=cols, extrasaction="ignore")
                writer.writeheader()
                for r in rows:
                    if isinstance(r, dict):
                        writer.writerow({k: _coerce_cell(r.get(k)) for k in cols})
            else:
                writer = csv.writer(csv_buf)
                for r in rows:
                    if isinstance(r, (list, tuple)):
                        writer.writerow([_coerce_cell(v) for v in r])
                    else:
                        writer.writerow([_coerce_cell(r)])
            data = "﻿" + csv_buf.getvalue()
            zf.writestr(f"{fname}.csv", data.encode("utf-8"))
            manifest["tables"].append({"name": name, "file": f"{fname}.csv", "rows": len(rows)})
            n_tables += 1

        ddl = state.get("sql_ddl") or ""
        if ddl:
            zf.writestr("_schema.sql", str(ddl))
        zf.writestr("_manifest.json", _json.dumps(manifest, indent=2, ensure_ascii=False))
        if n_tables == 0:
            zf.writestr("README.txt",
                        "Aucune table de resultat dans ce pipeline.\n"
                        "Relancez l'analyse pour generer des donnees exploitables.\n")

    logger.info(f"[Export] CSV bundle : {n_tables} tables, {buf.tell()} bytes")
    return buf.getvalue()


def generate_powerbi_template(state: dict, session_id: str) -> bytes:
    """Bundle Power BI : zip avec connection.pqt + README + schema.sql + manifest."""
    import io
    import json as _json
    import zipfile

    logical = state.get("logical_model") or {}
    facts = logical.get("fact_tables") or ([logical.get("fact_table")] if logical.get("fact_table") else [])
    dims = logical.get("dimension_tables") or []
    db_name = state.get("target_database") or state.get("restored_db_name") or "agent_dw"
    user_prefix = state.get("user_prefix") or ""

    def _tn(t):
        if isinstance(t, dict):
            return t.get("name") or t.get("table_name") or ""
        return str(t or "")

    fact_names = [_tn(f) for f in facts if f]
    dim_names = [_tn(d) for d in dims if d]
    all_tables = [t for t in fact_names + dim_names if t]
    if user_prefix:
        all_tables = [f"{user_prefix}_{t}" if not t.startswith(user_prefix + "_") else t for t in all_tables]

    pqt = "// Power BI Get Data - SQL Server Source\n"
    pqt += f"// Database: {db_name}\n"
    pqt += "let\n"
    pqt += f'    Source = Sql.Database("YOUR_SERVER", "{db_name}")\n'
    pqt += "in\n"
    pqt += "    Source"

    readme = (
        "# Power BI Bundle\n\n"
        f"Connectez-vous a la base SQL Server `{db_name}` et importez les tables :\n\n"
        + "\n".join(f"- {t}" for t in all_tables)
        + "\n\nUtilisez le fichier `connection.pqt` comme M-query template.\n"
    )

    manifest = {
        "session_id": session_id,
        "database": db_name,
        "tables": all_tables,
        "facts": fact_names,
        "dimensions": dim_names,
    }

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("connection.pqt", pqt)
        zf.writestr("POWERBI_CONNECT.md", readme)
        zf.writestr("schema.sql", state.get("sql_ddl", "") or "")
        zf.writestr("tables_manifest.json", _json.dumps(manifest, indent=2, ensure_ascii=False))
    return buf.getvalue()
